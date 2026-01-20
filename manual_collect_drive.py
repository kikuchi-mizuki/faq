#!/usr/bin/env python3
"""
Google Driveから文書を手動収集（詳細ログ付き）
"""

import os
import sys
import json
import base64
from dotenv import load_dotenv
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials

# 環境変数の読み込み
load_dotenv()

def manual_collect_drive():
    """Google Driveから文書を手動収集"""

    print("=" * 60)
    print("Google Drive 文書収集（手動）")
    print("=" * 60)

    # 認証情報の取得
    print("\n🔧 認証情報を取得中...")
    service_account_json_raw = os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON')

    if not service_account_json_raw:
        print("❌ GOOGLE_SERVICE_ACCOUNT_JSON環境変数が設定されていません")
        return

    # ファイルパスの場合は読み込み
    if service_account_json_raw.endswith('.json'):
        print(f"📄 JSONファイルから読み込み: {service_account_json_raw}")
        with open(service_account_json_raw, 'r') as f:
            credentials_info = json.load(f)
    else:
        # Base64デコード
        try:
            service_account_json = base64.b64decode(service_account_json_raw).decode('utf-8')
            credentials_info = json.loads(service_account_json)
            print("✅ Base64デコード成功")
        except Exception as e:
            print(f"❌ Base64デコード失敗: {e}")
            return

    # 認証
    credentials = Credentials.from_service_account_info(
        credentials_info,
        scopes=[
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive',
            'https://www.googleapis.com/auth/documents'
        ]
    )

    print(f"✅ 認証情報を取得しました")
    print(f"   サービスアカウント: {credentials_info.get('client_email')}")

    # Google Drive APIサービスを初期化
    print("\n🔧 Google Drive APIサービスを初期化中...")
    drive_service = build('drive', 'v3', credentials=credentials)
    print("✅ Google Drive APIサービスを初期化しました")

    # PDF、Excel、テキストファイルを検索
    print("\n📂 Google Driveファイルを検索中...")
    query = (
        "mimeType='text/plain' or "
        "mimeType='text/csv' or "
        "mimeType='application/pdf' or "
        "mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or "
        "mimeType='application/vnd.ms-excel'"
    )

    print(f"   クエリ: {query}")

    try:
        results = drive_service.files().list(
            q=query,
            fields="files(id, name, mimeType, modifiedTime)",
            pageSize=100
        ).execute()

        files = results.get('files', [])
        print(f"\n✅ {len(files)}件のファイルが見つかりました")

        if not files:
            print("\n⚠️ PDF/Excel/テキストファイルが見つかりません")
            print("\n💡 対処方法:")
            print("1. Google Driveにファイルをアップロード")
            print("2. サービスアカウントに共有権限を付与")
            print(f"   サービスアカウント: {credentials_info.get('client_email')}")
            return

        print("\n📄 ファイル一覧:")
        print("-" * 60)
        for i, file in enumerate(files, 1):
            print(f"\n{i}. 名前: {file['name']}")
            print(f"   ID: {file['id']}")
            print(f"   MimeType: {file['mimeType']}")
            print(f"   更新日時: {file['modifiedTime']}")

        # RAGサービスに追加
        print("\n" + "=" * 60)
        print("📚 RAGサービスにファイルを追加中...")
        print("-" * 60)

        sys.path.insert(0, os.path.dirname(__file__))
        from line_qa_system.rag_service import RAGService

        rag_service = RAGService()

        if not rag_service.is_enabled:
            print("❌ RAGサービスが有効になっていません")
            return

        print("✅ RAGサービスが有効です")

        # 各ファイルを処理
        for i, file in enumerate(files, 1):
            print(f"\n[{i}/{len(files)}] {file['name']} を処理中...")

            try:
                mime_type = file.get('mimeType', '')
                content = None

                # PDFファイルの処理
                if mime_type == 'application/pdf':
                    print(f"   📄 PDFファイルをダウンロード中...")
                    request = drive_service.files().get_media(fileId=file['id'])
                    pdf_content = request.execute()

                    # PyPDF2でテキスト抽出
                    try:
                        from PyPDF2 import PdfReader
                        import io
                        pdf_file = io.BytesIO(pdf_content)
                        pdf_reader = PdfReader(pdf_file)

                        text_parts = []
                        for page_num, page in enumerate(pdf_reader.pages, 1):
                            text = page.extract_text()
                            if text:
                                text_parts.append(f"=== ページ {page_num} ===\n{text}")

                        content = "\n\n".join(text_parts)
                        print(f"   ✅ {len(pdf_reader.pages)}ページのテキストを抽出しました")
                    except ImportError:
                        print(f"   ⚠️ PyPDF2がインストールされていません")
                        content = None

                # Excelファイルの処理
                elif mime_type in [
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'application/vnd.ms-excel'
                ]:
                    print(f"   📊 Excelファイルをダウンロード中...")
                    request = drive_service.files().get_media(fileId=file['id'])
                    excel_content = request.execute()

                    # openpyxlで読み込み
                    try:
                        import openpyxl
                        import io
                        excel_file = io.BytesIO(excel_content)
                        workbook = openpyxl.load_workbook(excel_file, data_only=True)

                        text_parts = []
                        for sheet_name in workbook.sheetnames:
                            worksheet = workbook[sheet_name]
                            sheet_text = [f"=== シート: {sheet_name} ==="]

                            # ヘッダー行の取得
                            headers = []
                            for cell in worksheet[1]:
                                if cell.value:
                                    headers.append(str(cell.value))

                            if headers:
                                sheet_text.append(f"列: {', '.join(headers)}")

                            # データ行の処理
                            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                                row_values = [str(val) if val is not None else "" for val in row]
                                if any(row_values):
                                    row_text = " | ".join([f"{h}={v}" for h, v in zip(headers, row_values) if v])
                                    if row_text:
                                        sheet_text.append(f"行{row_num}: {row_text}")

                            text_parts.append("\n".join(sheet_text))

                        content = "\n\n".join(text_parts)
                        print(f"   ✅ {len(workbook.sheetnames)}シートのデータを抽出しました")
                    except ImportError:
                        print(f"   ⚠️ openpyxlがインストールされていません")
                        content = None

                # テキストファイルの処理
                else:
                    print(f"   📝 テキストファイルをダウンロード中...")
                    request = drive_service.files().get_media(fileId=file['id'])
                    file_content = request.execute()

                    if isinstance(file_content, bytes):
                        content = file_content.decode('utf-8', errors='ignore')
                    else:
                        content = file_content

                    print(f"   ✅ テキストを取得しました")

                if content and content.strip():
                    rag_service.add_document(
                        source_type="google_drive",
                        source_id=file['id'],
                        title=file['name'],
                        content=content,
                        metadata={
                            "file_id": file['id'],
                            "mime_type": file['mimeType'],
                            "modified_time": file['modifiedTime'],
                            "collected_at": "manual_collection"
                        }
                    )
                    print(f"   ✅ RAGに追加しました（{len(content)}文字）")
                else:
                    print(f"   ⚠️ コンテンツが空でした")

            except Exception as e:
                print(f"   ❌ エラー: {e}")
                import traceback
                traceback.print_exc()

        print("\n" + "=" * 60)
        print("✅ 文書収集が完了しました")
        print("=" * 60)

    except Exception as e:
        print(f"\n❌ エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    manual_collect_drive()
