"""
文書収集サービス
Google Sheets、Google Docs、Google Driveから文書を収集
"""

import os
import json
import time
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
import structlog

import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# PDF/Excel解析用ライブラリ（条件付きインポート）
try:
    import pdfplumber
    PDF_SUPPORT = True
    PDF_LIBRARY = 'pdfplumber'
except ImportError:
    try:
        from PyPDF2 import PdfReader
        PDF_SUPPORT = True
        PDF_LIBRARY = 'pypdf2'
    except ImportError:
        PDF_SUPPORT = False
        PDF_LIBRARY = None
        logger = structlog.get_logger(__name__)
        logger.warning("PDFライブラリがインストールされていません。PDF解析機能が無効です。")

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    logger = structlog.get_logger(__name__)
    logger.warning("openpyxlがインストールされていません。Excel解析機能が無効です。")

from .config import Config
from .rag_service import RAGService

logger = structlog.get_logger(__name__)


class DocumentCollector:
    """文書収集サービス"""

    def __init__(self, rag_service: RAGService):
        """初期化"""
        self.rag_service = rag_service
        self.sheet_id = Config.SHEET_ID_QA

        # Google認証情報の取得
        self.credentials = self._get_credentials()
        self._drive_service = None
        self._docs_service = None

    def _get_credentials(self) -> Optional[Credentials]:
        """Google認証情報を取得"""
        try:
            service_account_json_raw = os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON')
            if not service_account_json_raw:
                logger.error("GOOGLE_SERVICE_ACCOUNT_JSON環境変数が設定されていません")
                return None
            
            logger.debug(f"Raw GOOGLE_SERVICE_ACCOUNT_JSON (first 50 chars): {service_account_json_raw[:50]}...")
            
            # Base64デコードを試行
            try:
                import base64
                service_account_json = base64.b64decode(service_account_json_raw).decode('utf-8')
                logger.debug(f"Base64 decoded successfully, first 50 chars: {service_account_json[:50]}...")
            except Exception as e:
                logger.warning(f"Base64デコードに失敗しました。生のJSONとして処理します: {e}")
                service_account_json = service_account_json_raw
            
            # JSON形式の検証
            try:
                credentials_info = json.loads(service_account_json)
            except json.JSONDecodeError as e:
                logger.error("GOOGLE_SERVICE_ACCOUNT_JSONの形式が正しくありません", error=str(e))
                logger.error(f"不正なJSON文字列の先頭: {service_account_json[:100]}")
                logger.info("RAG機能の文書収集は無効化されます。基本機能のみ利用可能です。")
                return None
            
            credentials = Credentials.from_service_account_info(
                credentials_info,
                scopes=[
                    'https://www.googleapis.com/auth/spreadsheets',
                    'https://www.googleapis.com/auth/drive',
                    'https://www.googleapis.com/auth/documents'
                ]
            )
            
            logger.info("Google認証情報を取得しました")
            return credentials
            
        except Exception as e:
            logger.error("Google認証情報の取得に失敗しました", error=str(e))
            logger.info("RAG機能の文書収集は無効化されます。基本機能のみ利用可能です。")
            return None

    @property
    def drive_service(self):
        """Google Drive APIサービス（遅延初期化）"""
        if self._drive_service is None and self.credentials:
            logger.info("Google Drive APIサービスを初期化しています...")
            self._drive_service = build('drive', 'v3', credentials=self.credentials)
            logger.info("Google Drive APIサービスの初期化が完了しました")
        return self._drive_service

    @property
    def docs_service(self):
        """Google Docs APIサービス（遅延初期化）"""
        if self._docs_service is None and self.credentials:
            logger.info("Google Docs APIサービスを初期化しています...")
            self._docs_service = build('docs', 'v1', credentials=self.credentials)
            logger.info("Google Docs APIサービスの初期化が完了しました")
        return self._docs_service

    def collect_all_documents(self) -> bool:
        """全ての文書を収集"""
        try:
            print("📚 文書収集を開始します")
            logger.info("文書収集を開始します")

            # Google Sheetsから文書を収集
            self._collect_sheets_documents()

            # Google Docsから文書を収集
            self._collect_docs_documents()

            # Google Driveから文書を収集
            self._collect_drive_documents()

            print("✅ 文書収集が完了しました")
            logger.info("文書収集が完了しました")
            return True

        except Exception as e:
            print(f"❌ 文書収集中にエラーが発生しました: {e}")
            logger.error("文書収集中にエラーが発生しました", error=str(e))
            return False

    def _collect_sheets_documents(self):
        """Google Sheetsから文書を収集"""
        try:
            gc = gspread.authorize(self.credentials)
            spreadsheet = gc.open_by_key(self.sheet_id)
            
            # 各シートを処理
            for worksheet in spreadsheet.worksheets():
                sheet_name = worksheet.title
                logger.info(f"シート '{sheet_name}' を処理中...")
                
                # シートのデータを取得
                data = worksheet.get_all_records()
                
                if not data:
                    continue
                
                # シートの内容をテキストに変換
                content = self._convert_sheet_to_text(data, sheet_name)
                
                if content:
                    # RAGサービスに追加
                    self.rag_service.add_document(
                        source_type="google_sheets",
                        source_id=f"{self.sheet_id}_{sheet_name}",
                        title=f"スプレッドシート: {sheet_name}",
                        content=content,
                        metadata={
                            "sheet_name": sheet_name,
                            "row_count": len(data),
                            "collected_at": datetime.now().isoformat()
                        }
                    )
                    
                    logger.info(f"シート '{sheet_name}' を収集しました")
            
        except Exception as e:
            logger.error("Google Sheets収集中にエラーが発生しました", error=str(e))

    def _collect_docs_documents(self):
        """Google Docsから文書を収集"""
        if not self.docs_service:
            logger.warning("Google Docsサービスが利用できません")
            return
        
        try:
            # Google Docsファイルを検索
            query = "mimeType='application/vnd.google-apps.document'"
            results = self.drive_service.files().list(
                q=query,
                fields="files(id, name, modifiedTime)"
            ).execute()
            
            files = results.get('files', [])
            logger.info(f"Google Docsファイルを{len(files)}件発見しました")
            
            for file in files:
                try:
                    # 文書の内容を取得
                    doc = self.docs_service.documents().get(documentId=file['id']).execute()
                    content = self._extract_docs_content(doc)
                    
                    if content:
                        # RAGサービスに追加
                        self.rag_service.add_document(
                            source_type="google_docs",
                            source_id=file['id'],
                            title=file['name'],
                            content=content,
                            metadata={
                                "file_id": file['id'],
                                "modified_time": file['modifiedTime'],
                                "collected_at": datetime.now().isoformat()
                            }
                        )
                        
                        logger.info(f"Google Docs '{file['name']}' を収集しました")
                
                except Exception as e:
                    logger.error(f"Google Docs '{file['name']}' の処理中にエラーが発生しました", error=str(e))
                    continue
            
        except Exception as e:
            logger.error("Google Docs収集中にエラーが発生しました", error=str(e))

    def _collect_drive_documents(self):
        """Google Driveから文書を収集"""
        if not self.drive_service:
            logger.warning("Google Driveサービスが利用できません")
            return

        try:
            # PDF、Excel、テキストファイルを検索
            query = (
                "mimeType='text/plain' or "
                "mimeType='text/csv' or "
                "mimeType='application/pdf' or "
                "mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or "
                "mimeType='application/vnd.ms-excel'"
            )
            results = self.drive_service.files().list(
                q=query,
                fields="files(id, name, mimeType, modifiedTime)"
            ).execute()

            files = results.get('files', [])
            print(f"📁 Google Driveファイルを{len(files)}件発見しました")
            logger.info(f"Google Driveファイルを{len(files)}件発見しました")

            for file in files:
                try:
                    # ファイルの内容を取得
                    content = self._extract_drive_file_content(file)

                    if content:
                        # RAGサービスに追加
                        self.rag_service.add_document(
                            source_type="google_drive",
                            source_id=file['id'],
                            title=file['name'],
                            content=content,
                            metadata={
                                "file_id": file['id'],
                                "mime_type": file['mimeType'],
                                "modified_time": file['modifiedTime'],
                                "collected_at": datetime.now().isoformat()
                            }
                        )

                        print(f"✅ Google Driveファイル '{file['name']}' を収集しました")
                        logger.info(f"Google Driveファイル '{file['name']}' を収集しました")

                except Exception as e:
                    logger.error(f"Google Driveファイル '{file['name']}' の処理中にエラーが発生しました", error=str(e))
                    continue

        except Exception as e:
            logger.error("Google Drive収集中にエラーが発生しました", error=str(e))

    def _convert_sheet_to_text(self, data: List[Dict[str, Any]], sheet_name: str) -> str:
        """シートデータをテキストに変換"""
        if not data:
            return ""
        
        # ヘッダーを取得
        headers = list(data[0].keys()) if data else []
        
        # テキストを構築
        text_parts = [f"シート名: {sheet_name}"]
        text_parts.append(f"列: {', '.join(headers)}")
        text_parts.append("")
        
        # 各行を処理
        for i, row in enumerate(data, 1):
            row_text = f"行{i}: "
            row_parts = []
            
            for header in headers:
                value = row.get(header, '')
                if value:
                    row_parts.append(f"{header}={value}")
            
            if row_parts:
                row_text += " | ".join(row_parts)
                text_parts.append(row_text)
        
        return "\n".join(text_parts)

    def _extract_docs_content(self, doc: Dict[str, Any]) -> str:
        """Google Docsの内容を抽出"""
        try:
            content_parts = []
            
            # 文書の構造を解析
            if 'body' in doc and 'content' in doc['body']:
                for element in doc['body']['content']:
                    if 'paragraph' in element:
                        paragraph = element['paragraph']
                        if 'elements' in paragraph:
                            for elem in paragraph['elements']:
                                if 'textRun' in elem and 'content' in elem['textRun']:
                                    content_parts.append(elem['textRun']['content'])
            
            return ''.join(content_parts).strip()
            
        except Exception as e:
            logger.error("Google Docs内容抽出中にエラーが発生しました", error=str(e))
            return ""

    def _extract_drive_file_content(self, file: Dict[str, Any]) -> str:
        """Google Driveファイルの内容を抽出"""
        if not self.drive_service:
            logger.error(f"Google Driveサービスが初期化されていません: {file['name']}")
            return ""

        try:
            mime_type = file.get('mimeType', '')

            # PDFファイルの処理
            if mime_type == 'application/pdf':
                return self._extract_pdf_content(file)

            # Excelファイルの処理
            elif mime_type in [
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'application/vnd.ms-excel'
            ]:
                return self._extract_excel_content(file)

            # テキストファイルの処理
            else:
                request = self.drive_service.files().get_media(fileId=file['id'])
                content = request.execute()

                # テキストとしてデコード
                if isinstance(content, bytes):
                    content = content.decode('utf-8', errors='ignore')

                return content.strip()

        except Exception as e:
            logger.error(f"Google Driveファイル内容抽出中にエラーが発生しました: {file['name']}", error=str(e))
            return ""

    def _extract_pdf_content(self, file: Dict[str, Any]) -> str:
        """PDFファイルからテキストを抽出"""
        if not PDF_SUPPORT:
            logger.warning(f"PDF解析がサポートされていません: {file['name']}")
            return ""

        if not self.drive_service:
            logger.error(f"Google Driveサービスが初期化されていません: {file['name']}")
            return ""

        try:
            # PDFファイルをダウンロード
            request = self.drive_service.files().get_media(fileId=file['id'])
            pdf_content = request.execute()
            pdf_file = io.BytesIO(pdf_content)

            text_parts = []

            # pdfplumberを優先的に使用（日本語対応が優れている）
            if PDF_LIBRARY == 'pdfplumber':
                print(f"📖 pdfplumberを使用してPDFを解析します: {file['name']}")
                logger.info(f"pdfplumberを使用してPDFを解析します: {file['name']}")
                with pdfplumber.open(pdf_file) as pdf:
                    for page_num, page in enumerate(pdf.pages, 1):
                        try:
                            text = page.extract_text()
                            if text:
                                text_parts.append(f"=== ページ {page_num} ===\n{text}")
                        except Exception as e:
                            logger.warning(f"PDFページ {page_num} の抽出に失敗: {file['name']}", error=str(e))
                            continue
                    print(f"✅ PDFから{len(pdf.pages)}ページのテキストを抽出しました: {file['name']}")
                    logger.info(f"PDFから{len(pdf.pages)}ページのテキストを抽出しました: {file['name']}")

            # PyPDF2をフォールバックとして使用
            else:
                logger.info(f"PyPDF2を使用してPDFを解析します: {file['name']}")
                from PyPDF2 import PdfReader
                pdf_reader = PdfReader(pdf_file)
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    try:
                        text = page.extract_text()
                        if text:
                            text_parts.append(f"=== ページ {page_num} ===\n{text}")
                    except Exception as e:
                        logger.warning(f"PDFページ {page_num} の抽出に失敗: {file['name']}", error=str(e))
                        continue
                logger.info(f"PDFから{len(pdf_reader.pages)}ページのテキストを抽出しました: {file['name']}")

            extracted_text = "\n\n".join(text_parts)
            return extracted_text

        except Exception as e:
            logger.error(f"PDF内容抽出中にエラーが発生しました: {file['name']}", error=str(e))
            return ""

    def _extract_excel_content(self, file: Dict[str, Any]) -> str:
        """Excelファイルからテキストを抽出"""
        if not EXCEL_SUPPORT:
            logger.warning(f"Excel解析がサポートされていません: {file['name']}")
            return ""

        if not self.drive_service:
            logger.error(f"Google Driveサービスが初期化されていません: {file['name']}")
            return ""

        try:
            # Excelファイルをダウンロード
            request = self.drive_service.files().get_media(fileId=file['id'])
            excel_content = request.execute()

            # openpyxlで読み込み
            excel_file = io.BytesIO(excel_content)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)

            text_parts = []
            for sheet_name in workbook.sheetnames:
                try:
                    worksheet = workbook[sheet_name]
                    sheet_text = [f"=== シート: {sheet_name} ==="]

                    # ヘッダー行の取得
                    headers = []
                    for cell in worksheet[1]:
                        if cell.value:
                            headers.append(str(cell.value))

                    if headers:
                        sheet_text.append(f"列: {', '.join(headers)}")

                    # データ行の処理
                    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                        row_values = [str(val) if val is not None else "" for val in row]
                        if any(row_values):  # 空行をスキップ
                            row_text = " | ".join([f"{h}={v}" for h, v in zip(headers, row_values) if v])
                            if row_text:
                                sheet_text.append(f"行{row_num}: {row_text}")

                    text_parts.append("\n".join(sheet_text))

                except Exception as e:
                    logger.warning(f"Excelシート '{sheet_name}' の処理に失敗: {file['name']}", error=str(e))
                    continue

            extracted_text = "\n\n".join(text_parts)
            print(f"📊 Excelから{len(workbook.sheetnames)}シートのデータを抽出しました: {file['name']}")
            logger.info(f"Excelから{len(workbook.sheetnames)}シートのデータを抽出しました: {file['name']}")
            return extracted_text

        except Exception as e:
            logger.error(f"Excel内容抽出中にエラーが発生しました: {file['name']}", error=str(e))
            return ""

    def health_check(self) -> bool:
        """文書収集サービスの健全性チェック"""
        return self.credentials is not None and self.rag_service.health_check()
