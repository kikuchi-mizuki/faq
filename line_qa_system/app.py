"""
LINE Q&A自動応答システム - メインアプリケーション
"""

import os
import time
import hashlib
import hmac
import base64
import json
import threading
from datetime import datetime
from typing import Dict, Any, Optional
from functools import wraps
from contextlib import contextmanager

import structlog
from flask import Flask, request, jsonify, abort
from dotenv import load_dotenv

from .line_client import LineClient
from .qa_service import QAService
from .session_service import SessionService
from .flow_service import FlowService
from .rag_service import RAGService
from .document_collector import DocumentCollector
from .config import Config
from .utils import verify_line_signature, hash_user_id

# 環境変数の読み込み
load_dotenv()

# 構造化ログの設定
structlog.configure(
    processors=[
        structlog.stdlib.filter_by_level,
        structlog.stdlib.add_logger_name,
        structlog.stdlib.add_log_level,
        structlog.stdlib.PositionalArgumentsFormatter(),
        structlog.processors.TimeStamper(fmt="iso"),
        structlog.processors.StackInfoRenderer(),
        structlog.processors.format_exc_info,
        structlog.processors.UnicodeDecoder(),
        structlog.processors.JSONRenderer(),
    ],
    context_class=dict,
    logger_factory=structlog.stdlib.LoggerFactory(),
    wrapper_class=structlog.stdlib.BoundLogger,
    cache_logger_on_first_use=True,
)

logger = structlog.get_logger(__name__)

app = Flask(__name__)
app.config.from_object(Config)

# Flaskの最大ファイルサイズを設定
app.config['MAX_CONTENT_LENGTH'] = Config.MAX_FILE_SIZE_MB * 1024 * 1024

# サービスの初期化（遅延初期化）
qa_service = None
line_client = None
session_service = None
flow_service = None
rag_service = None
document_collector = None

# レート制限用のインメモリキャッシュ（IPアドレス -> アップロード時刻のリスト）
upload_rate_limiter = {}

def check_upload_rate_limit(ip_address: str) -> bool:
    """
    アップロードのレート制限をチェック

    Args:
        ip_address: クライアントのIPアドレス

    Returns:
        True: アップロード可能, False: レート制限超過
    """
    current_time = time.time()
    hour_ago = current_time - 3600  # 1時間前

    # 古いエントリをクリーンアップ
    if ip_address in upload_rate_limiter:
        upload_rate_limiter[ip_address] = [
            t for t in upload_rate_limiter[ip_address] if t > hour_ago
        ]
    else:
        upload_rate_limiter[ip_address] = []

    # レート制限チェック
    if len(upload_rate_limiter[ip_address]) >= Config.UPLOAD_RATE_LIMIT_PER_HOUR:
        logger.warning(f"レート制限超過: IP={ip_address}, count={len(upload_rate_limiter[ip_address])}")
        return False

    # アップロード時刻を記録
    upload_rate_limiter[ip_address].append(current_time)
    return True

def safe_error_message(error: Exception, default_message: str = "処理中にエラーが発生しました") -> str:
    """
    本番環境では汎用的なエラーメッセージを返し、開発環境では詳細を返す

    Args:
        error: 例外オブジェクト
        default_message: デフォルトのエラーメッセージ

    Returns:
        エラーメッセージ文字列
    """
    if Config.is_production():
        # 本番環境では汎用メッセージのみ（詳細はログに記録）
        logger.error(f"エラー詳細（本番環境）: {str(error)}", exc_info=True)
        return default_message
    else:
        # 開発環境では詳細なエラー情報を返す
        return f"{default_message}: {str(error)}"

@contextmanager
def rag_db_connection():
    """
    RAGデータベース接続のコンテキストマネージャ

    使用例:
        with rag_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute("SELECT ...")

    Raises:
        ValueError: RAGサービスが無効な場合
        ConnectionError: データベース接続の取得に失敗した場合
    """
    if not rag_service or not rag_service.is_enabled or not rag_service.db_pool:
        raise ValueError("RAGサービスが無効です")

    conn = rag_service.get_db_connection()
    if not conn:
        raise ConnectionError("データベース接続の取得に失敗しました")

    try:
        yield conn
    finally:
        rag_service.return_db_connection(conn)

def initialize_services():
    """サービスの初期化（遅延初期化）"""
    global qa_service, line_client, session_service, flow_service, rag_service, document_collector

    if qa_service is not None:
        print("✅ サービスは既に初期化済みです")
        return  # 既に初期化済み

    try:
        print("🚀 サービスの初期化を開始します...")
        logger.info("サービスの初期化を開始します")

        # 環境変数の検証
        validation_errors = Config.validate()
        if validation_errors:
            logger.error("環境変数の検証エラー", errors=validation_errors)
            for error in validation_errors:
                print(f"❌ {error}")

            # 本番環境では致命的エラーとして起動を中止
            if Config.is_production():
                raise ValueError(f"環境変数の検証に失敗しました: {', '.join(validation_errors)}")
            else:
                # 開発環境では警告のみ
                print("⚠️ 開発環境のため、警告として続行します")
                logger.warning("開発環境のため、検証エラーを無視して続行します")

        # 本番環境のセキュリティチェック
        Config.check_production_security()

        # AIServiceの初期化（最優先で行い、他サービスへ注入）
        from .ai_service import AIService
        ai_service = AIService()
        logger.info(f"AIServiceの初期化完了: is_enabled={ai_service.is_enabled}")
        
        # AIサービスが無効な場合の詳細ログ
        if not ai_service.is_enabled:
            logger.warning("AIサービスが無効です。GEMINI_API_KEYの設定を確認してください。")
        else:
            logger.info("AIサービスが有効です。")
        
        # QAService（AIServiceを渡す）
        qa_service = QAService(ai_service)
        logger.info("QAServiceの初期化が完了しました")
        
        line_client = LineClient()
        logger.info("LineClientの初期化が完了しました")
        
        session_service = SessionService()
        logger.info("SessionServiceの初期化が完了しました")
        
        # RAGサービスの初期化（段階的有効化）
        rag_service = None
        try:
            logger.info("RAGServiceの初期化を開始します")
            rag_service = RAGService()
            logger.info(f"RAGServiceの初期化完了: is_enabled={rag_service.is_enabled}")
        except Exception as e:
            logger.error("RAG機能の初期化に失敗しました", error=str(e), exc_info=True)
            logger.info("RAG機能は無効化されています。基本機能のみ利用可能です。")
            rag_service = None  # 明示的にNoneを設定
        
        flow_service = FlowService(session_service, qa_service, rag_service, ai_service)
        logger.info("FlowServiceの初期化が完了しました")
        
        # DocumentCollectorの初期化（RAG機能が有効な場合）
        document_collector = None
        if rag_service and rag_service.is_enabled:
            try:
                document_collector = DocumentCollector(rag_service)
                print("📄 DocumentCollectorの初期化が完了しました")
                logger.info("DocumentCollectorの初期化が完了しました")

                # 起動時の自動文書収集を無効化（手動トリガーのみ有効）
                # 理由: Railway起動時のタイムアウトとメモリ制約を考慮
                print("📚 自動文書収集は無効化されています（/admin/collect-documents で手動実行可能）")
                logger.info("自動文書収集は無効化されています。管理APIで手動実行してください。")

                # 定期的な自動収集も無効化
                # start_auto_document_collection()

            except Exception as e:
                logger.error("DocumentCollectorの初期化に失敗しました", error=str(e), exc_info=True)
                document_collector = None  # 明示的にNoneを設定
        else:
            logger.info("RAGサービスが無効なため、DocumentCollectorは初期化されません")

        print("✅ 全てのサービスの初期化が完了しました")
        logger.info("全てのサービスの初期化が完了しました")

    except Exception as e:
        import traceback
        print(f"❌❌❌ サービスの初期化中にエラーが発生しました: {e}")
        print("=== エラーの詳細（Traceback） ===")
        print(traceback.format_exc())
        logger.error("サービスの初期化中にエラーが発生しました", error=str(e), exc_info=True)
        # エラーが発生してもアプリケーションは起動する
        logger.warning("一部のサービスが初期化できませんでした。基本機能のみ利用可能です。")


def start_auto_document_collection():
    """定期的な文書収集を開始（1時間ごと）"""
    def auto_collect_worker():
        while True:
            try:
                time.sleep(3600)  # 1時間ごと
                logger.info("定期文書収集を開始します")

                # DocumentCollectorが初期化されている場合のみ実行
                if document_collector:
                    try:
                        success = document_collector.collect_all_documents()
                        if success:
                            logger.info("✅ 定期文書収集が完了しました")
                        else:
                            logger.warning("⚠️ 定期文書収集でエラーが発生しました")
                    except Exception as e:
                        logger.error("定期文書収集中にエラーが発生しました", error=str(e))
                else:
                    logger.warning("DocumentCollectorが初期化されていないため、定期文書収集をスキップします")

            except Exception as e:
                logger.error("定期文書収集ワーカーでエラーが発生しました", error=str(e))

    # バックグラウンドで定期文書収集を開始
    collect_thread = threading.Thread(target=auto_collect_worker, daemon=True)
    collect_thread.start()
    logger.info("定期文書収集機能を開始しました（1時間ごと）")


def start_auto_reload():
    """定期的な自動リロード機能を開始"""
    last_sheet_update = None

    def auto_reload_worker():
        nonlocal last_sheet_update
        while True:
            try:
                time.sleep(900)  # 15分ごと（API制限を考慮）
                logger.info("自動リロードチェック開始")

                # スプレッドシートの最終更新時刻をチェック
                try:
                    # 現在の最終更新時刻を取得（簡易実装）
                    current_time = time.time()

                    # サービスが初期化されていない場合はスキップ
                    if qa_service is None or flow_service is None:
                        logger.info("サービスが初期化されていないため、自動リロードをスキップします")
                        continue

                    # 初回実行時または強制リロード時
                    if last_sheet_update is None:
                        qa_service.reload_cache()
                        last_sheet_update = current_time
                        logger.info("初回自動リロード完了")
                    else:
                        # 通常の定期リロード（変更検知なし）
                        qa_service.reload_cache()
                        flow_service.reload_flows()
                        logger.info("定期自動リロード完了")

                except Exception as e:
                    logger.error("スプレッドシート更新チェック中にエラー", error=str(e))

            except Exception as e:
                logger.error("自動リロード中にエラーが発生しました", error=str(e))

    # バックグラウンドで自動リロードを開始
    reload_thread = threading.Thread(target=auto_reload_worker, daemon=True)
    reload_thread.start()
    logger.info("自動リロード機能を開始しました")


# アプリケーション起動時に自動リロードを開始
start_auto_reload()


def require_admin(f):
    """管理者権限が必要なエンドポイント用デコレータ（APIキー認証）"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        api_key = request.headers.get("X-API-Key")
        expected_api_key = app.config.get("ADMIN_API_KEY", "")

        # APIキーが設定されていない場合はエラー
        if not expected_api_key:
            logger.error("ADMIN_API_KEYが設定されていません")
            abort(500, description="サーバー設定エラー: ADMIN_API_KEYが未設定です")

        # APIキーの検証
        if not api_key:
            logger.warning("APIキーが提供されていません")
            abort(401, description="認証が必要です。X-API-Keyヘッダーを含めてください")

        if api_key != expected_api_key:
            logger.warning("無効なAPIキーが提供されました")
            abort(403, description="無効なAPIキーです")

        logger.info("管理者認証成功")
        return f(*args, **kwargs)

    return decorated_function


@app.route("/callback", methods=["POST"])
def callback():
    """LINE Webhook受信エンドポイント"""
    start_time = time.time()

    try:
        # リクエストボディを一度だけ取得（複数回読むとエラーになるため）
        body_bytes = request.get_data()

        # LINE署名の検証
        signature = request.headers.get("X-Line-Signature", "")
        if not verify_line_signature(signature, body_bytes, app.config["LINE_CHANNEL_SECRET"]):
            logger.warning("LINE署名検証に失敗しました",
                         signature=signature[:20] if signature else "なし")
            # 署名検証失敗の場合、200を返すが処理はスキップ（セキュリティ対策）
            return jsonify({"status": "ok"})

        # リクエストボディの解析
        try:
            body = json.loads(body_bytes.decode('utf-8'))
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            logger.warning("リクエストボディの解析に失敗しました", error=str(e))
            # 不正なボディでも200を返す
            return jsonify({"status": "ok"})

        if not body:
            logger.warning("リクエストボディが空です")
            return jsonify({"status": "ok"})

        # イベントの処理
        events = body.get("events", [])
        print(f"📥 Webhook受信: {len(events)}個のイベント")
        logger.info(f"Webhook受信: {len(events)}個のイベント")

        for event in events:
            print(f"📋 イベントタイプ: {event.get('type')}")
            logger.info(f"イベント処理", event_type=event.get('type'))

            if event["type"] == "message" and event["message"]["type"] == "text":
                print(f"✅ テキストメッセージを処理します")
                process_text_message(event, start_time)
            elif event["type"] == "postback":
                print(f"✅ ポストバックを処理します")
                process_postback_message(event, start_time)
            else:
                print(f"⚠️ 未対応のイベントタイプ: {event.get('type')}")

        return jsonify({"status": "ok"})

    except Exception as e:
        logger.error("Webhook処理中にエラーが発生しました", error=str(e), exc_info=True)
        # エラーが発生しても200を返す（LINEの要件）
        return jsonify({"status": "ok"})


def process_postback_message(event: Dict[str, Any], start_time: float):
    """ポストバックメッセージの処理"""
    user_id = event["source"]["userId"]
    reply_token = event["replyToken"]
    
    # ユーザーIDのハッシュ化
    hashed_user_id = hash_user_id(user_id)
    
    logger.info("ポストバックを受信しました", user_id=hashed_user_id)
    
    try:
        # 認証フローの処理
        if Config.AUTH_ENABLED:
            from .auth_flow import AuthFlow
            auth_flow = AuthFlow()
            
            if auth_flow.handle_postback(event):
                return  # 認証フローで処理された場合は終了
        
        # その他のポストバック処理
        # 必要に応じて追加
        
    except Exception as e:
        logger.error("ポストバック処理中にエラーが発生しました", 
                    user_id=hashed_user_id, 
                    error=str(e), 
                    exc_info=True)
        
        # エラー時の応答
        try:
            line_client.reply_text(reply_token, "申し訳ございません。一時的なエラーが発生しました。")
        except Exception as reply_error:
            logger.error("エラーメッセージの送信に失敗しました", error=str(reply_error))


def process_text_message(event: Dict[str, Any], start_time: float):
    """テキストメッセージの処理（認証チェック付き）"""
    user_id = event["source"]["userId"]
    message_text = event["message"]["text"]
    reply_token = event["replyToken"]

    # ユーザーIDのハッシュ化
    hashed_user_id = hash_user_id(user_id)

    print(f"📨 メッセージを受信: user={hashed_user_id}, text='{message_text}'")
    logger.info("メッセージを受信しました", user_id=hashed_user_id, text=message_text)

    try:
        print("🔧 メッセージ処理を開始します")
        # サービスが初期化されていない場合は初期化を試行
        if qa_service is None or line_client is None:
            initialize_services()
        
        # サービスが初期化されていない場合はエラーメッセージを返す
        if qa_service is None or line_client is None:
            try:
                # 最低限のLineClientを初期化
                temp_line_client = LineClient()
                temp_line_client.reply_text(reply_token, "申し訳ございません。システムの初期化中です。しばらくお待ちください。")
            except Exception as e:
                logger.error("エラーメッセージの送信に失敗しました", error=str(e))
            return
        
        # 認証チェック（認証が有効な場合）
        logger.info("認証機能の状態確認", AUTH_ENABLED=Config.AUTH_ENABLED)

        if Config.AUTH_ENABLED:
            logger.info("認証チェックを開始します", user_id=hashed_user_id)
            from .optimized_auth_flow import OptimizedAuthFlow
            auth_flow = OptimizedAuthFlow()

            # 認証フローの処理
            if auth_flow.process_auth_flow(event):
                logger.info("認証フローで処理されました", user_id=hashed_user_id)
                return  # 認証フローで処理された場合は終了

            # 認証済みでない場合は制限メッセージを送信
            try:
                is_authenticated = auth_flow.is_authenticated(user_id)
                logger.info("認証チェック結果",
                           user_id=hashed_user_id,
                           is_authenticated=is_authenticated)

                if not is_authenticated:
                    auth_flow.send_auth_required_message(reply_token)
                    logger.info("未認証ユーザーをブロックしました", user_id=hashed_user_id)
                    return
                else:
                    logger.info("認証済みユーザーを許可しました", user_id=hashed_user_id)
            except Exception as e:
                logger.error("認証チェック中にエラーが発生しました",
                           user_id=hashed_user_id,
                           error=str(e))
                # エラーが発生した場合は安全のため認証が必要メッセージを送信
                try:
                    auth_flow.send_auth_required_message(reply_token)
                except:
                    pass
                return
        else:
            logger.info("認証機能が無効です。誰でもアクセスできます", user_id=hashed_user_id)
        # キャンセルコマンドのチェック
        if message_text.strip().lower() in ["キャンセル", "cancel", "やめる", "終了"]:
            if flow_service.is_in_flow(user_id):
                flow_service.cancel_flow(user_id)
                line_client.reply_text(reply_token, "会話をキャンセルしました。")
                logger.info("フローをキャンセルしました", user_id=hashed_user_id)
                return
            else:
                line_client.reply_text(reply_token, "現在、会話は進行していません。")
                return

        # フロー中かどうかをチェック
        if flow_service.is_in_flow(user_id):
            # フロー中の場合は選択を処理
            next_flow, is_end = flow_service.process_user_choice(user_id, message_text)

            if next_flow:
                # 次のステップがある場合
                if is_end:
                    # 終了ステップ（回答）
                    line_client.reply_text(reply_token, next_flow.question)
                    logger.info("フロー終了", user_id=hashed_user_id)
                else:
                    # 次の質問を提示（クイックリプライ付き）
                    options = next_flow.option_list
                    line_client.reply_text(
                        reply_token, next_flow.question, quick_reply=options if options else None
                    )
                    logger.info(
                        "次のステップへ進みました", user_id=hashed_user_id, step=next_flow.step
                    )
            else:
                # フローが見つからない場合
                line_client.reply_text(
                    reply_token,
                    "申し訳ございません。処理中にエラーが発生しました。最初からやり直してください。",
                )
                flow_service.cancel_flow(user_id)
                logger.warning("フロー処理に失敗しました", user_id=hashed_user_id)

        else:
            # フロー外の場合は通常のQ&A検索
            # まず、AI文脈判断でフローのトリガーかどうかをチェック
            flow = flow_service.find_flow_by_ai_context(message_text)
            if flow:
                # AI文脈判断でフローを開始
                flow_service.start_flow(user_id, flow.trigger)
                # 最初の質問を送信（クイックリプライ付き）
                options = flow.option_list
                line_client.reply_text(
                    reply_token,
                    flow.question,
                    quick_reply=options if options else None,
                )
                logger.info(
                    "AI文脈判断でフローを開始しました", user_id=hashed_user_id, trigger=flow.trigger
                )
                return
            
            # 自然言語マッチングも試行
            flow = flow_service.find_flow_by_natural_language(message_text)
            if flow:
                # 自然言語マッチングでフローを開始
                flow_service.start_flow(user_id, flow.trigger)
                # 最初の質問を送信（クイックリプライ付き）
                options = flow.option_list
                line_client.reply_text(
                    reply_token,
                    flow.question,
                    quick_reply=options if options else None,
                )
                logger.info(
                    "自然言語マッチングでフローを開始しました", user_id=hashed_user_id, trigger=flow.trigger
                )
                return
            
            # 従来の厳密マッチングも試行
            available_triggers = flow_service.get_available_triggers()
            for trigger in available_triggers:
                if trigger.lower() in message_text.lower():
                    # フローを開始
                    flow = flow_service.start_flow(user_id, trigger)
                    if flow:
                        # 最初の質問を送信（クイックリプライ付き）
                        options = flow.option_list
                        line_client.reply_text(
                            reply_token,
                            flow.question,
                            quick_reply=options if options else None,
                        )
                        logger.info(
                            "フローを開始しました", user_id=hashed_user_id, trigger=trigger
                        )
                        return

            # フローに該当しない場合は通常のQ&A検索
            result = qa_service.find_answer(message_text)

            # 認証情報の取得（ログ用）
            store_code = ""
            staff_id = ""
            if Config.AUTH_ENABLED:
                try:
                    from .optimized_auth_flow import OptimizedAuthFlow
                    auth_flow = OptimizedAuthFlow()
                    auth_data = auth_flow.get_auth_info(user_id)
                    if auth_data:
                        store_code = auth_data.get('store_code', '')
                        staff_id = auth_data.get('staff_id', '')
                        logger.debug("認証情報を取得しました",
                                   user_id=hash_user_id(user_id),
                                   store_code=store_code,
                                   staff_id=staff_id)
                except Exception as e:
                    logger.error("認証情報の取得に失敗しました", error=str(e))

            # 質問をログに記録
            qa_service.log_query(
                user_id_hash=hashed_user_id,
                query=message_text,
                result=result,
                store_code=store_code,
                staff_id=staff_id
            )

            # 応答の送信
            if result.is_found and result.top_result is not None:
                response_text = format_answer(
                    result.top_result.answer,
                    result.top_result.question,
                    getattr(result.top_result, "tags", "")
                )
                line_client.reply_text(reply_token, response_text)

                logger.info(
                    "回答を送信しました",
                    user_id=hashed_user_id,
                    question_id=getattr(result.top_result, "id", None),
                    score=getattr(result.top_result, "score", None),
                )
            else:
                # 明確なマッチが見つからない場合、RAGで回答を試行
                rag_answer = None
                print(f"🤔 Q&Aに該当なし。RAGで回答を試行します: message='{message_text}'")
                logger.info("Q&Aに該当なし。RAGで回答を試行", user_id=hashed_user_id, message=message_text)

                # 処理中メッセージを送信（RAG検索は時間がかかるため）
                if rag_service and rag_service.is_enabled:
                    try:
                        line_client.push_message(user_id, "💭 考え中です...")
                        print(f"💬 処理中メッセージを送信しました")
                        logger.info("処理中メッセージを送信", user_id=hashed_user_id)
                    except Exception as e:
                        print(f"⚠️ 処理中メッセージの送信に失敗: {e}")
                        logger.warning("処理中メッセージの送信に失敗", error=str(e))

                if rag_service and rag_service.is_enabled:
                    print(f"✅ RAGサービスが有効です")
                    try:
                        # RAGで類似文書を検索
                        print(f"🔍 RAGで類似文書を検索しています...")
                        similar_docs = rag_service.search_similar_documents(message_text, limit=10)
                        print(f"🔍 検索結果: {len(similar_docs)}件の類似文書")

                        if similar_docs:
                            # 類似文書が見つかった場合、コンテキストを構築してAI回答生成
                            print(f"✅ 類似文書が見つかりました。AI回答を生成します")
                            context = rag_service._build_context(similar_docs)
                            rag_answer = rag_service.generate_answer(message_text, context)
                            print(f"✅ RAG回答生成完了")
                            logger.info("RAGで回答を生成しました", user_id=hashed_user_id, doc_count=len(similar_docs))
                        else:
                            print(f"⚠️ RAGで類似文書が見つかりませんでした")
                            logger.info("RAGで類似文書が見つかりませんでした", user_id=hashed_user_id)
                    except Exception as e:
                        print(f"❌ RAG回答生成中にエラー: {e}")
                        logger.error("RAG回答生成中にエラーが発生しました", user_id=hashed_user_id, error=str(e))
                else:
                    print(f"❌ RAGサービスが無効です: rag_service={rag_service is not None}, is_enabled={rag_service.is_enabled if rag_service else 'N/A'}")

                # RAGで回答が得られた場合はそれを返す、そうでなければエスカレーション
                if rag_answer:
                    response_text = f"{rag_answer}\n\n※この回答はアップロードされた資料から生成されました。"
                    line_client.reply_text(reply_token, response_text)
                    logger.info("RAG回答を送信しました", user_id=hashed_user_id)
                else:
                    # RAGでも回答が得られない場合は本社スタッフへエスカレーション
                    fallback_text = get_fallback_response()
                    line_client.reply_text(reply_token, fallback_text)
                    logger.info("該当なし - 本社スタッフへエスカレーション", user_id=hashed_user_id)

        # 処理時間の記録
        latency = int((time.time() - start_time) * 1000)
        logger.info("メッセージ処理完了", user_id=hashed_user_id, latency_ms=latency)

    except Exception as e:
        logger.error(
            "メッセージ処理中にエラーが発生しました", user_id=hashed_user_id, error=str(e), exc_info=True
        )

        # エラー時の応答
        try:
            error_text = "申し訳ございません。一時的なエラーが発生しました。"
            line_client.reply_text(reply_token, error_text)
        except Exception as reply_error:
            logger.error("エラーメッセージの送信にも失敗しました", error=str(reply_error))


def format_answer(answer: str, question: str, tags: str) -> str:
    """回答のフォーマット"""
    # シンプルにanswerだけを返す
    return answer


def format_candidates(candidates: list) -> str:
    """候補のフォーマット"""
    text = "もしかしてこれですか？\n\n"
    for i, candidate in enumerate(candidates[:3], 1):
        tags_text = f" ({candidate.tags})" if candidate.tags else ""
        text += f"{i}. {candidate.question}{tags_text}\n"
    text += "\nより具体的なキーワードをお試しください。"
    return text




def get_fallback_response() -> str:
    """フォールバック応答"""
    return "こちらの質問は本社スタッフまでお願いします！"


@app.route("/healthz", methods=["GET"])
def health_check():
    """ヘルスチェックエンドポイント"""
    try:
        print("=== ヘルスチェック開始 ===")
        # サービスが初期化されていない場合は初期化を試行
        if qa_service is None:
            print("⚠️ サービスが未初期化です。初期化を試行します...")
            logger.info("ヘルスチェック時にサービスを初期化します")
            initialize_services()
            print(f"✅ 初期化完了: qa_service={qa_service is not None}")
        else:
            print("✅ サービスは既に初期化済みです")
            logger.info("サービスは既に初期化済みです")

        # 基本的な健全性チェック（QAが生きていればOK、他は情報として返す）
        print(f"🔍 ヘルスチェック実行中: qa_service={qa_service is not None}")
        qa_healthy = qa_service.health_check() if qa_service is not None else False
        flow_loaded = (flow_service is not None and len(flow_service.flows) > 0)
        ai_healthy = (flow_service is not None and flow_service.ai_service.health_check()) if flow_service is not None else False

        print(f"📊 ヘルスチェック結果: qa_healthy={qa_healthy}, flow_loaded={flow_loaded}, ai_healthy={ai_healthy}")

        if qa_healthy:
            print("✅ ヘルスチェック成功")

            # RAG診断情報を追加（debug=trueパラメータの場合のみ）
            response_data = {
                "status": "healthy",
                "timestamp": time.time(),
                "version": "0.1.0",
                "qa_service": "ok",
                "flow_service_loaded": flow_loaded,
                "ai_service": "ok" if ai_healthy else "disabled"
            }

            # debug=true の場合、詳細なRAG診断情報を含める
            if request.args.get('debug') == 'true':
                try:
                    rag_diagnostic = {
                        "rag_service_initialized": rag_service is not None,
                        "rag_service_enabled": rag_service.is_enabled if rag_service else False,
                        "db_connected": rag_service.db_pool is not None if rag_service else False,
                        "embedding_model_loaded": rag_service.embedding_model is not None if rag_service else False,
                        "gemini_api_key_set": bool(os.getenv('GEMINI_API_KEY')),
                        "database_url_set": bool(os.getenv('DATABASE_URL')),
                        "rag_lightweight_mode": os.getenv('RAG_LIGHTWEIGHT_MODE', 'false'),
                        "similarity_threshold": os.getenv('SIMILARITY_THRESHOLD', '0.15'),
                    }

                    # 文書数とEmbedding数をカウント
                    if rag_service and rag_service.db_pool:
                        try:
                            conn = rag_service.get_db_connection()
                            with conn.cursor() as cursor:
                                cursor.execute("SELECT COUNT(*) FROM documents WHERE chunk_index >= 0")
                                doc_count = cursor.fetchone()[0]
                                cursor.execute("SELECT COUNT(*) FROM document_embeddings")
                                embedding_count = cursor.fetchone()[0]
                                rag_diagnostic["document_count"] = doc_count
                                rag_diagnostic["embedding_count"] = embedding_count
                            rag_service.return_db_connection(conn)
                        except Exception as db_error:
                            rag_diagnostic["db_error"] = str(db_error)

                    response_data["rag_diagnostic"] = rag_diagnostic
                except Exception as diag_error:
                    response_data["rag_diagnostic_error"] = str(diag_error)

            return jsonify(response_data)
        else:
            print("❌ ヘルスチェック失敗: QAサービスが不健全")
            return jsonify({
                "status": "unhealthy",
                "qa_service": "error",
                "flow_service_loaded": flow_loaded,
                "ai_service": "ok" if ai_healthy else "disabled",
                "timestamp": time.time()
            }), 500

    except Exception as e:
        import traceback
        print(f"❌ ヘルスチェックで例外発生: {e}")
        print(traceback.format_exc())
        logger.error("ヘルスチェックに失敗しました", error=str(e), exc_info=True)
        return (
            jsonify({"status": "unhealthy", "error": str(e), "timestamp": time.time()}),
            500,
        )


@app.route("/admin/reload", methods=["POST"])
@require_admin
def reload_cache():
    """キャッシュの再読み込み（管理者のみ）"""
    try:
        qa_service.reload_cache()
        flow_service.reload_flows()
        logger.info("手動リロードが完了しました")
        return jsonify({
            "status": "success",
            "message": "キャッシュを再読み込みしました（Q&A + フロー + 資料）",
            "timestamp": time.time(),
            "auto_reload_active": True
        })
    except Exception as e:
        logger.error("キャッシュの再読み込みに失敗しました", error=str(e))
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/admin/rag-status", methods=["GET"])
@require_admin
def rag_status():
    """RAG機能の状態を確認（管理者のみ）"""
    try:
        # データベース内の文書数を確認
        document_count = 0
        embedding_count = 0
        db_connected = False
        embedding_model_loaded = False

        if rag_service and rag_service.is_enabled:
            db_connected = rag_service.db_connection is not None
            embedding_model_loaded = rag_service.embedding_model is not None

            if rag_service.db_connection:
                try:
                    with rag_service.db_connection.cursor() as cursor:
                        cursor.execute("SELECT COUNT(*) FROM documents;")
                        document_count = cursor.fetchone()[0]

                        cursor.execute("SELECT COUNT(*) FROM document_embeddings;")
                        embedding_count = cursor.fetchone()[0]
                except Exception as db_error:
                    logger.error("DB文書数の取得に失敗", error=str(db_error))

        return jsonify({
            "status": "success",
            "rag_service_initialized": rag_service is not None,
            "rag_service_enabled": rag_service.is_enabled if rag_service else False,
            "db_connected": db_connected,
            "embedding_model_loaded": embedding_model_loaded,
            "document_count": document_count,
            "embedding_count": embedding_count,
            "document_collector_initialized": document_collector is not None,
            "gemini_api_key_set": bool(os.getenv('GEMINI_API_KEY')),
            "database_url_set": bool(os.getenv('DATABASE_URL')),
            "timestamp": time.time()
        })
    except Exception as e:
        logger.error("RAG状態確認に失敗しました", error=str(e))
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/admin/documents", methods=["GET"])
@require_admin
def list_documents():
    """登録されている文書の一覧を取得（管理者のみ）"""
    conn = None
    try:
        if not rag_service or not rag_service.is_enabled or not rag_service.db_pool:
            return jsonify({
                "status": "error",
                "message": "RAGサービスまたはDB接続が無効です"
            }), 500

        # 接続プールから接続を取得
        conn = rag_service.get_db_connection()
        if not conn:
            logger.error("データベース接続の取得に失敗しました")
            return jsonify({
                "status": "error",
                "message": safe_error_message(None, "データベース接続の取得に失敗しました")
            }), 500

        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT
                    source_type,
                    source_id,
                    title,
                    COUNT(*) as chunk_count,
                    MAX(created_at) as last_updated
                FROM documents
                WHERE chunk_index >= 0
                GROUP BY source_type, source_id, title
                ORDER BY last_updated DESC
                LIMIT 50;
            """)
            results = cursor.fetchall()

            documents = []
            for row in results:
                documents.append({
                    "source_type": row[0],
                    "source_id": row[1],
                    "title": row[2],
                    "chunk_count": row[3],
                    "last_updated": str(row[4])
                })

            return jsonify({
                "status": "success",
                "total_documents": len(documents),
                "documents": documents
            })

    except Exception as e:
        logger.error("文書一覧の取得に失敗しました", error=str(e))
        return jsonify({
            "status": "error",
            "message": safe_error_message(e, "文書一覧の取得に失敗しました")
        }), 500
    finally:
        if conn:
            rag_service.return_db_connection(conn)


@app.route("/documents", methods=["GET"])
def list_documents_public():
    """登録されている文書の一覧を取得（公開エンドポイント）"""
    conn = None
    try:
        if not rag_service or not rag_service.is_enabled or not rag_service.db_pool:
            return jsonify({
                "status": "error",
                "message": "RAGサービスまたはDB接続が無効です"
            }), 503

        # 接続プールから接続を取得
        conn = rag_service.get_db_connection()
        if not conn:
            logger.error("データベース接続の取得に失敗しました")
            return jsonify({
                "status": "error",
                "message": "データベース接続の取得に失敗しました"
            }), 500

        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT
                    d.source_type,
                    d.source_id,
                    d.title,
                    COUNT(DISTINCT d.id) as chunk_count,
                    MAX(d.created_at) as last_updated,
                    COUNT(DISTINCT e.document_id) > 0 as has_embeddings
                FROM documents d
                LEFT JOIN document_embeddings e ON d.id = e.document_id
                WHERE d.chunk_index >= 0
                GROUP BY d.source_type, d.source_id, d.title
                ORDER BY last_updated DESC
                LIMIT 100;
            """)
            results = cursor.fetchall()

            documents = []
            for row in results:
                documents.append({
                    "source_type": row[0],
                    "source_id": row[1],
                    "title": row[2],
                    "chunk_count": row[3],
                    "last_updated": str(row[4]),
                    "has_embeddings": bool(row[5])
                })

            return jsonify({
                "status": "success",
                "total_documents": len(documents),
                "documents": documents
            })

    except Exception as e:
        logger.error("文書一覧の取得に失敗しました", error=str(e))
        return jsonify({
            "status": "error",
            "message": safe_error_message(e, "文書一覧の取得に失敗しました")
        }), 500
    finally:
        if conn:
            rag_service.return_db_connection(conn)


@app.route("/delete-document/<source_id>", methods=["DELETE"])
def delete_document(source_id):
    """文書を削除（公開エンドポイント）"""
    # 入力検証: source_id
    import re
    if not re.match(r'^[a-zA-Z0-9_-]+$', source_id):
        return jsonify({
            "status": "error",
            "message": "無効なsource_idです"
        }), 400

    # 入力検証: source_type
    source_type = request.args.get('source_type', 'upload')
    logger.info(f"削除リクエスト: source_id={source_id}, source_type={source_type}, allowed={Config.ALLOWED_SOURCE_TYPES}")
    if source_type not in Config.ALLOWED_SOURCE_TYPES:
        logger.error(f"無効なsource_type: {source_type} (allowed: {Config.ALLOWED_SOURCE_TYPES})")
        return jsonify({
            "status": "error",
            "message": f"無効なsource_typeです: {source_type}"
        }), 400

    conn = None
    try:
        # クライアントIPアドレスを取得
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        if ',' in client_ip:
            client_ip = client_ip.split(',')[0].strip()

        # レート制限チェック（削除もアップロードと同じレート制限を使用）
        if not check_upload_rate_limit(client_ip):
            logger.warning(f"削除レート制限超過: IP={client_ip}")
            return jsonify({
                "status": "error",
                "message": f"操作回数の上限に達しました。1時間あたり{Config.UPLOAD_RATE_LIMIT_PER_HOUR}回までです。"
            }), 429

        # RAGサービスの状態確認
        if not rag_service or not rag_service.is_enabled or not rag_service.db_pool:
            return jsonify({
                "status": "error",
                "message": "RAGサービスが無効です"
            }), 503

        # 接続プールから接続を取得
        conn = rag_service.get_db_connection()
        if not conn:
            logger.error("データベース接続の取得に失敗しました")
            return jsonify({
                "status": "error",
                "message": "データベース接続の取得に失敗しました"
            }), 500

        # データベースから削除
        with conn.cursor() as cursor:
            # 削除対象の文書IDを取得
            if source_type:
                cursor.execute("""
                    SELECT id FROM documents
                    WHERE source_id = %s AND source_type = %s;
                """, (source_id, source_type))
            else:
                cursor.execute("""
                    SELECT id FROM documents
                    WHERE source_id = %s;
                """, (source_id,))

            doc_ids = [row[0] for row in cursor.fetchall()]

            if not doc_ids:
                return jsonify({
                    "status": "error",
                    "message": "指定された文書が見つかりません"
                }), 404

            # Embeddingを削除
            cursor.execute("""
                DELETE FROM document_embeddings
                WHERE document_id = ANY(%s);
            """, (doc_ids,))
            embedding_deleted = cursor.rowcount

            # 文書を削除
            if source_type:
                cursor.execute("""
                    DELETE FROM documents
                    WHERE source_id = %s AND source_type = %s;
                """, (source_id, source_type))
            else:
                cursor.execute("""
                    DELETE FROM documents
                    WHERE source_id = %s;
                """, (source_id,))

            doc_deleted = cursor.rowcount
            conn.commit()

            logger.info(f"文書を削除しました: source_id={source_id}, docs={doc_deleted}, embeddings={embedding_deleted}")

            return jsonify({
                "status": "success",
                "message": f"文書を削除しました（{doc_deleted}件の文書、{embedding_deleted}件のEmbedding）",
                "deleted_documents": doc_deleted,
                "deleted_embeddings": embedding_deleted
            })

    except Exception as e:
        logger.error("文書削除に失敗しました", error=str(e), exc_info=True)
        # ロールバック
        if conn:
            try:
                conn.rollback()
            except:
                pass
        return jsonify({
            "status": "error",
            "message": safe_error_message(e, "文書削除に失敗しました")
        }), 500
    finally:
        if conn:
            rag_service.return_db_connection(conn)


@app.route("/download-document/<source_id>", methods=["GET"])
def download_document(source_id):
    """文書をダウンロード（公開エンドポイント）"""
    # 入力検証: source_id
    import re
    if not re.match(r'^[a-zA-Z0-9_-]+$', source_id):
        return jsonify({
            "status": "error",
            "message": "無効なsource_idです"
        }), 400

    # 入力検証: source_type
    source_type = request.args.get('source_type', 'upload')
    logger.info(f"ダウンロードリクエスト: source_id={source_id}, source_type={source_type}, allowed={Config.ALLOWED_SOURCE_TYPES}")
    if source_type not in Config.ALLOWED_SOURCE_TYPES:
        logger.error(f"無効なsource_type: {source_type} (allowed: {Config.ALLOWED_SOURCE_TYPES})")
        return jsonify({
            "status": "error",
            "message": f"無効なsource_typeです: {source_type}"
        }), 400

    conn = None
    try:
        # RAGサービスの状態確認
        if not rag_service or not rag_service.is_enabled or not rag_service.db_pool:
            return jsonify({
                "status": "error",
                "message": "RAGサービスが無効です"
            }), 503

        # 接続プールから接続を取得
        conn = rag_service.get_db_connection()
        if not conn:
            logger.error("データベース接続の取得に失敗しました")
            return jsonify({
                "status": "error",
                "message": "データベース接続の取得に失敗しました"
            }), 500

        # データベースから文書を取得（全チャンクを結合）
        with conn.cursor() as cursor:
            # chunk_index >= 0のチャンクを順番に取得
            if source_type:
                cursor.execute("""
                    SELECT title, content, chunk_index, full_content
                    FROM documents
                    WHERE source_id = %s AND source_type = %s AND chunk_index >= 0
                    ORDER BY chunk_index ASC;
                """, (source_id, source_type))
            else:
                cursor.execute("""
                    SELECT title, content, chunk_index, full_content
                    FROM documents
                    WHERE source_id = %s AND chunk_index >= 0
                    ORDER BY chunk_index ASC;
                """, (source_id,))

            results = cursor.fetchall()

            if not results:
                return jsonify({
                    "status": "error",
                    "message": "指定された文書が見つかりません"
                }), 404

            # タイトルを取得
            title = results[0][0]

            # full_contentがあればそれを使用、なければチャンクを結合
            full_content = results[0][3]  # full_content

            if not full_content:
                # full_contentがない場合、チャンクを結合
                content_parts = [row[1] for row in results]  # content
                full_content = "\n\n".join(content_parts)

            # ファイル名を決定（拡張子を保持）
            filename = title if title else f"document_{source_id}.txt"

            # Content-Typeを決定
            content_type = "text/plain; charset=utf-8"
            if filename.endswith('.pdf'):
                content_type = "application/pdf"
            elif filename.endswith(('.xlsx', '.xls')):
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            # レスポンスを返す
            from flask import Response
            from urllib.parse import quote

            # 日本語ファイル名をRFC 5987形式でエンコード
            encoded_filename = quote(filename)

            response = Response(
                full_content,
                mimetype=content_type,
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                    "Content-Type": content_type
                }
            )

            logger.info(f"文書をダウンロード: source_id={source_id}, title={title}")
            return response

    except Exception as e:
        logger.error("文書ダウンロードに失敗しました", error=str(e), exc_info=True)
        return jsonify({
            "status": "error",
            "message": safe_error_message(e, "文書ダウンロードに失敗しました")
        }), 500
    finally:
        if conn:
            rag_service.return_db_connection(conn)


@app.route("/generate-embeddings", methods=["POST"])
def generate_embeddings_for_pending():
    """Embedding未生成の文書に対してEmbeddingを生成（誰でもアクセス可能）"""
    conn = None
    try:
        if not rag_service or not rag_service.is_enabled or not rag_service.db_pool:
            return jsonify({
                "status": "error",
                "message": "RAGサービスが無効です"
            }), 503

        # 接続プールから接続を取得
        conn = rag_service.get_db_connection()
        if not conn:
            return jsonify({
                "status": "error",
                "message": "データベース接続の取得に失敗しました"
            }), 500

        # Embedding未生成の文書を検索（全文レコードは除外）
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT d.id, d.content
                FROM documents d
                LEFT JOIN document_embeddings e ON d.id = e.document_id
                WHERE e.document_id IS NULL
                  AND d.chunk_index >= 0
                LIMIT 100;
            """)
            pending_docs = cursor.fetchall()

        if not pending_docs:
            return jsonify({
                "status": "success",
                "message": "Embedding生成が必要な文書はありません",
                "generated_count": 0
            })

        # Embeddingを生成
        generated_count = 0
        for doc_id, content in pending_docs:
            try:
                embedding = rag_service._generate_embedding(content)
                embedding_str = '[' + ','.join(map(str, embedding.tolist())) + ']'

                with conn.cursor() as cursor:
                    cursor.execute(
                        "INSERT INTO document_embeddings (document_id, embedding) VALUES (%s, %s::vector)",
                        (doc_id, embedding_str)
                    )
                conn.commit()
                generated_count += 1
            except Exception as e:
                logger.error(f"Embedding生成エラー (doc_id={doc_id})", error=str(e))
                continue

        return jsonify({
            "status": "success",
            "message": f"{generated_count}個の文書のEmbeddingを生成しました",
            "generated_count": generated_count
        })

    except Exception as e:
        logger.error("Embedding生成に失敗しました", error=str(e), exc_info=True)
        return jsonify({
            "status": "error",
            "message": safe_error_message(e, "Embedding生成に失敗しました")
        }), 500
    finally:
        if conn:
            rag_service.return_db_connection(conn)


@app.route("/generate-embeddings/<source_id>", methods=["POST"])
def generate_embeddings_for_source(source_id):
    """特定のsource_idに対してEmbeddingを生成（誰でもアクセス可能）"""
    conn = None
    try:
        # 入力検証: source_id
        import re
        if not re.match(r'^[a-zA-Z0-9_-]+$', source_id):
            return jsonify({"status": "error", "message": "無効なsource_idです"}), 400

        source_type = request.args.get('source_type', 'upload')
        if source_type not in Config.ALLOWED_SOURCE_TYPES:
            return jsonify({"status": "error", "message": "無効なsource_typeです"}), 400

        if not rag_service or not rag_service.is_enabled or not rag_service.db_pool:
            return jsonify({
                "status": "error",
                "message": "RAGサービスが無効です"
            }), 503

        # 接続プールから接続を取得
        conn = rag_service.get_db_connection()
        if not conn:
            return jsonify({
                "status": "error",
                "message": "データベース接続の取得に失敗しました"
            }), 500

        # 指定されたsource_idのEmbedding未生成の文書を検索
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT d.id, d.content
                FROM documents d
                LEFT JOIN document_embeddings e ON d.id = e.document_id
                WHERE e.document_id IS NULL
                  AND d.source_id = %s
                  AND d.source_type = %s
                  AND d.chunk_index >= 0;
            """, (source_id, source_type))
            pending_docs = cursor.fetchall()

        if not pending_docs:
            return jsonify({
                "status": "success",
                "message": "このファイルのEmbeddingは既に生成済みです",
                "generated_count": 0
            })

        # Embeddingを生成
        generated_count = 0
        for doc_id, content in pending_docs:
            try:
                embedding = rag_service._generate_embedding(content)
                embedding_str = '[' + ','.join(map(str, embedding.tolist())) + ']'

                with conn.cursor() as cursor:
                    cursor.execute(
                        "INSERT INTO document_embeddings (document_id, embedding) VALUES (%s, %s::vector)",
                        (doc_id, embedding_str)
                    )
                conn.commit()
                generated_count += 1
            except Exception as e:
                logger.error(f"Embedding生成エラー (doc_id={doc_id})", error=str(e))
                continue

        return jsonify({
            "status": "success",
            "message": f"{generated_count}個のチャンクのEmbeddingを生成しました",
            "generated_count": generated_count
        })

    except Exception as e:
        logger.error("Embedding生成に失敗しました", error=str(e), exc_info=True)
        return jsonify({
            "status": "error",
            "message": safe_error_message(e, "Embedding生成に失敗しました")
        }), 500
    finally:
        if conn:
            rag_service.return_db_connection(conn)


@app.route("/admin/collect-documents", methods=["POST"])
@require_admin
def collect_documents():
    """Google Driveから文書を収集（管理者のみ）- 同期実行"""
    try:
        # RAGサービスの状態確認
        print("🔍 文書収集APIが呼ばれました")
        logger.info("文書収集APIが呼ばれました")

        if not rag_service or not rag_service.is_enabled:
            print("❌ RAGサービスが無効です")
            return jsonify({
                "status": "error",
                "message": "RAGサービスが無効です。GEMINI_API_KEYとDATABASE_URLを設定してください。",
                "rag_service_initialized": rag_service is not None,
                "rag_service_enabled": rag_service.is_enabled if rag_service else False
            }), 500

        if not document_collector:
            print("❌ DocumentCollectorが初期化されていません")
            return jsonify({
                "status": "error",
                "message": "DocumentCollectorが初期化されていません。RAG機能が無効の可能性があります。"
            }), 500

        print("✅ 文書収集を開始します（同期実行）")
        logger.info("文書収集を開始します")

        # 同期的に実行（デバッグ用）
        try:
            print("📝 document_collector.collect_all_documents()を呼び出します")
            success = document_collector.collect_all_documents()
            print(f"📝 文書収集結果: {success}")

            if success:
                logger.info("管理者による文書収集が完了しました")
                return jsonify({
                    "status": "success",
                    "message": "文書収集が完了しました",
                    "timestamp": time.time()
                })
            else:
                logger.error("文書収集中にエラーが発生しました")
                return jsonify({
                    "status": "error",
                    "message": "文書収集中にエラーが発生しました"
                }), 500

        except Exception as e:
            print(f"❌ 文書収集エラー: {e}")
            import traceback
            traceback.print_exc()
            logger.error("文書収集に失敗しました", error=str(e), exc_info=True)
            return jsonify({
                "status": "error",
                "message": safe_error_message(e, "文書収集中にエラーが発生しました")
            }), 500

    except Exception as e:
        print(f"❌ APIエラー: {e}")
        logger.error("文書収集APIでエラーが発生しました", error=str(e), exc_info=True)
        return jsonify({
            "status": "error",
            "message": safe_error_message(e, "文書収集APIでエラーが発生しました")
        }), 500


@app.route("/upload", methods=["GET"])
def upload_form():
    """ファイルアップロード・管理画面（誰でもアクセス可能）- Gemスタイル"""
    html = """
<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
    <meta http-equiv="Pragma" content="no-cache">
    <meta http-equiv="Expires" content="0">
    <title>ファイルアップロード</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Google Sans', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            background: #f8f9fa;
            min-height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 20px;
        }
        .container {
            max-width: 680px;
            width: 100%;
        }
        .header {
            text-align: center;
            margin-bottom: 32px;
        }
        .logo {
            width: 48px;
            height: 48px;
            background: linear-gradient(135deg, #8E44AD 0%, #3498DB 100%);
            border-radius: 50%;
            margin: 0 auto 16px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 24px;
        }
        h1 {
            font-size: 24px;
            font-weight: 400;
            color: #202124;
            margin-bottom: 8px;
        }
        .subtitle {
            font-size: 14px;
            color: #5f6368;
        }
        .card {
            background: white;
            border-radius: 12px;
            border: 1px solid #e8eaed;
            overflow: hidden;
            margin-bottom: 16px;
        }
        .tabs {
            display: flex;
            border-bottom: 1px solid #e8eaed;
            background: #f8f9fa;
        }
        .tab {
            flex: 1;
            padding: 16px;
            background: transparent;
            border: none;
            cursor: pointer;
            font-size: 14px;
            font-weight: 500;
            color: #5f6368;
            transition: all 0.2s;
            border-bottom: 2px solid transparent;
        }
        .tab.active {
            color: #1a73e8;
            border-bottom-color: #1a73e8;
            background: white;
        }
        .tab:hover:not(.active) {
            background: #f1f3f4;
        }
        .tab-content {
            display: none;
            padding: 32px;
        }
        .tab-content.active {
            display: block;
        }

        /* ドラッグ&ドロップエリア */
        .upload-area {
            border: 2px dashed #dadce0;
            border-radius: 8px;
            padding: 48px 24px;
            text-align: center;
            transition: all 0.3s;
            cursor: pointer;
            background: #f8f9fa;
            margin-bottom: 24px;
        }
        .upload-area.dragover {
            border-color: #1a73e8;
            background: #e8f0fe;
        }
        .upload-area:hover {
            border-color: #1a73e8;
            background: #f1f3f4;
        }
        .upload-icon {
            font-size: 48px;
            margin-bottom: 16px;
            opacity: 0.7;
        }
        .upload-text {
            font-size: 16px;
            color: #202124;
            margin-bottom: 8px;
        }
        .upload-hint {
            font-size: 13px;
            color: #5f6368;
        }
        .file-input {
            display: none;
        }

        /* 選択されたファイル表示 */
        .selected-file {
            display: none;
            background: #e8f0fe;
            border: 1px solid #1a73e8;
            border-radius: 8px;
            padding: 16px;
            margin-bottom: 24px;
            align-items: center;
            gap: 12px;
        }
        .selected-file.show {
            display: flex;
        }
        .file-icon {
            font-size: 32px;
        }
        .file-details {
            flex: 1;
        }
        .file-name {
            font-size: 14px;
            font-weight: 500;
            color: #202124;
            margin-bottom: 4px;
        }
        .file-size {
            font-size: 12px;
            color: #5f6368;
        }
        .remove-file {
            background: transparent;
            border: none;
            cursor: pointer;
            color: #5f6368;
            font-size: 20px;
            padding: 8px;
            border-radius: 50%;
            transition: background 0.2s;
        }
        .remove-file:hover {
            background: #f1f3f4;
        }

        /* 入力フィールド */
        .form-group {
            margin-bottom: 24px;
        }
        label {
            display: block;
            font-size: 13px;
            color: #5f6368;
            margin-bottom: 8px;
            font-weight: 500;
        }
        input[type="text"] {
            width: 100%;
            padding: 12px 16px;
            border: 1px solid #dadce0;
            border-radius: 8px;
            font-size: 14px;
            transition: all 0.2s;
            font-family: inherit;
        }
        input[type="text"]:focus {
            outline: none;
            border-color: #1a73e8;
            box-shadow: 0 0 0 4px rgba(26, 115, 232, 0.1);
        }

        /* ボタン */
        .btn {
            padding: 12px 24px;
            border: none;
            border-radius: 24px;
            font-size: 14px;
            font-weight: 500;
            cursor: pointer;
            transition: all 0.2s;
            font-family: inherit;
        }
        .btn-primary {
            background: #1a73e8;
            color: white;
        }
        .btn-primary:hover:not(:disabled) {
            background: #1557b0;
            box-shadow: 0 1px 2px 0 rgba(60,64,67,.3), 0 1px 3px 1px rgba(60,64,67,.15);
        }
        .btn-primary:disabled {
            background: #dadce0;
            cursor: not-allowed;
        }
        .btn-secondary {
            background: #f1f3f4;
            color: #5f6368;
        }
        .btn-secondary:hover {
            background: #e8eaed;
        }
        .btn-danger {
            background: #ea4335;
            color: white;
        }
        .btn-danger:hover {
            background: #d33b2c;
        }
        .btn-group {
            display: flex;
            gap: 12px;
            justify-content: flex-end;
        }

        /* メッセージ */
        .message {
            padding: 12px 16px;
            border-radius: 8px;
            font-size: 13px;
            margin-bottom: 24px;
            display: none;
            align-items: center;
            gap: 12px;
        }
        .message.show {
            display: flex;
        }
        .message.success {
            background: #e6f4ea;
            color: #137333;
            border: 1px solid #c6e1c6;
        }
        .message.error {
            background: #fce8e6;
            color: #c5221f;
            border: 1px solid #f4c7c3;
        }
        .message-icon {
            font-size: 20px;
        }

        /* ローダー */
        .loader {
            display: none;
            text-align: center;
            padding: 32px;
        }
        .loader.show {
            display: block;
        }
        .spinner {
            border: 3px solid #f1f3f4;
            border-top: 3px solid #1a73e8;
            border-radius: 50%;
            width: 40px;
            height: 40px;
            animation: spin 0.8s linear infinite;
            margin: 0 auto 16px;
        }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        .loader-text {
            font-size: 14px;
            color: #5f6368;
        }

        /* ファイルリスト */
        .file-list {
            margin-top: 24px;
        }
        .file-item {
            background: #f8f9fa;
            border: 1px solid #e8eaed;
            border-radius: 8px;
            padding: 16px;
            margin-bottom: 12px;
            display: flex;
            align-items: center;
            gap: 16px;
            transition: all 0.2s;
        }
        .file-item:hover {
            box-shadow: 0 1px 2px 0 rgba(60,64,67,.3), 0 1px 3px 1px rgba(60,64,67,.15);
        }
        .file-item-icon {
            font-size: 32px;
        }
        .file-item-info {
            flex: 1;
        }
        .file-item-title {
            font-size: 14px;
            font-weight: 500;
            color: #202124;
            margin-bottom: 4px;
        }
        .file-item-meta {
            font-size: 12px;
            color: #5f6368;
        }
        .file-item-badge {
            display: inline-block;
            padding: 2px 8px;
            background: #e8f0fe;
            color: #1a73e8;
            border-radius: 4px;
            font-size: 11px;
            margin-right: 8px;
            font-weight: 500;
        }
        .file-item-actions {
            display: flex;
            gap: 8px;
        }
        .icon-btn {
            background: transparent;
            border: none;
            cursor: pointer;
            padding: 8px;
            border-radius: 50%;
            transition: background 0.2s;
            font-size: 20px;
        }
        .icon-btn:hover {
            background: #f1f3f4;
        }
        .empty-state {
            text-align: center;
            padding: 48px 24px;
        }
        .empty-icon {
            font-size: 64px;
            opacity: 0.3;
            margin-bottom: 16px;
        }
        .empty-text {
            font-size: 14px;
            color: #5f6368;
        }
        .info-box {
            background: #f8f9fa;
            border-left: 4px solid #1a73e8;
            padding: 12px 16px;
            border-radius: 4px;
            font-size: 13px;
            color: #5f6368;
            margin-top: 16px;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="logo">📁</div>
            <h1>ファイルアップロード</h1>
            <p class="subtitle">ドキュメントをアップロードしてAIに学習させる</p>
        </div>

        <div class="card">
            <div class="tabs">
                <button class="tab active" onclick="switchTab('upload')">アップロード</button>
                <button class="tab" onclick="switchTab('list')">ファイル一覧</button>
            </div>

            <!-- アップロードタブ -->
            <div id="upload-tab" class="tab-content active">
                <div class="upload-area" id="dropArea">
                    <div class="upload-icon">📁</div>
                    <div class="upload-text">ファイルをドラッグ&ドロップ</div>
                    <div class="upload-hint">または クリックしてファイルを選択</div>
                </div>

                <div class="selected-file" id="selectedFile">
                    <div class="file-icon">📄</div>
                    <div class="file-details">
                        <div class="file-name" id="fileName"></div>
                        <div class="file-size" id="fileSize"></div>
                    </div>
                    <button class="remove-file" onclick="removeFile()">✕</button>
                </div>

                <input type="file" id="fileInput" class="file-input" accept=".pdf,.xlsx,.xls,.txt">

                <form id="uploadForm">
                    <div class="form-group">
                        <label for="title">タイトル（オプション）</label>
                        <input type="text" id="title" placeholder="このファイルで何をするかを入力してください">
                    </div>

                    <div class="info-box">
                        💡 対応形式: PDF, Excel (.xlsx, .xls), テキスト (.txt) | 最大サイズ: 5MB
                    </div>

                    <div id="uploadMessage" class="message"></div>

                    <div class="btn-group" style="margin-top: 24px;">
                        <button type="submit" id="uploadBtn" class="btn btn-primary" disabled>アップロード</button>
                    </div>
                </form>

                <div id="uploadLoader" class="loader">
                    <div class="spinner"></div>
                    <div class="loader-text">アップロード中...</div>
                </div>
            </div>

            <!-- ファイル一覧タブ -->
            <div id="list-tab" class="tab-content">
                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
                    <h2 style="font-size: 18px; font-weight: 500; color: #202124; margin: 0;">アップロード済みファイル</h2>
                    <div style="display: flex; gap: 8px;">
                        <button onclick="generateAllEmbeddings()" class="btn btn-primary" style="padding: 8px 16px; font-size: 13px;">🔮 全てのEmbedding生成</button>
                        <button onclick="loadDocuments()" class="btn btn-secondary" style="padding: 8px 16px; font-size: 13px;">🔄 更新</button>
                    </div>
                </div>

                <div id="listLoader" class="loader">
                    <div class="spinner"></div>
                    <div class="loader-text">読み込み中...</div>
                </div>

                <div id="listMessage" class="message"></div>

                <div id="fileList" class="file-list"></div>
            </div>
        </div>
    </div>

    <script>
        let selectedFile = null;

        // タブ切り替え（グローバルスコープに定義）
        function switchTab(tabName) {
            // 全てのタブとコンテンツからactiveクラスを削除
            document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));

            // 選択されたタブにactiveクラスを追加
            const tabs = document.querySelectorAll('.tab');
            tabs.forEach(tab => {
                if (tab.textContent.includes(tabName === 'upload' ? 'アップロード' : 'ファイル一覧')) {
                    tab.classList.add('active');
                }
            });

            // 対応するコンテンツにactiveクラスを追加
            document.getElementById(tabName + '-tab').classList.add('active');

            // ファイル一覧タブの場合はファイルを読み込む
            if (tabName === 'list') {
                loadDocuments();
            }
        }

        // DOMContentLoadedイベントで初期化
        document.addEventListener('DOMContentLoaded', function() {
            // DOM要素
            const dropArea = document.getElementById('dropArea');
            const fileInput = document.getElementById('fileInput');
            const selectedFileDiv = document.getElementById('selectedFile');
            const uploadBtn = document.getElementById('uploadBtn');
            const uploadForm = document.getElementById('uploadForm');

            // ドラッグ&ドロップイベント
            dropArea.addEventListener('click', () => fileInput.click());

        dropArea.addEventListener('dragover', (e) => {
            e.preventDefault();
            dropArea.classList.add('dragover');
        });

        dropArea.addEventListener('dragleave', () => {
            dropArea.classList.remove('dragover');
        });

        dropArea.addEventListener('drop', (e) => {
            e.preventDefault();
            dropArea.classList.remove('dragover');
            const files = e.dataTransfer.files;
            if (files.length > 0) {
                handleFileSelectInternal(files[0]);
            }
        });

        fileInput.addEventListener('change', (e) => {
            if (e.target.files.length > 0) {
                handleFileSelectInternal(e.target.files[0]);
            }
        });

            function handleFileSelectInternal(file) {
                selectedFile = file;
                const fileSize = (file.size / 1024).toFixed(1);
                const fileIcon = getFileIcon(file.name);

                document.getElementById('fileName').textContent = file.name;
                document.getElementById('fileSize').textContent = `${fileSize} KB`;
                document.querySelector('.file-icon').textContent = fileIcon;

                selectedFileDiv.classList.add('show');
                uploadBtn.disabled = false;

                // FileInputに同じファイルを設定
                const dataTransfer = new DataTransfer();
                dataTransfer.items.add(file);
                fileInput.files = dataTransfer.files;
            }

        // アップロード処理
        uploadForm.addEventListener('submit', async (e) => {
            e.preventDefault();

            if (!selectedFile) {
                showMessage('uploadMessage', 'error', '⚠ ファイルを選択してください');
                return;
            }

            const loader = document.getElementById('uploadLoader');
            const uploadBtn = document.getElementById('uploadBtn');

            uploadBtn.disabled = true;
            loader.classList.add('show');
            hideMessage('uploadMessage');

            const formData = new FormData();
            formData.append('file', selectedFile);
            const title = document.getElementById('title').value;
            if (title) {
                formData.append('title', title);
            }

            try {
                const response = await fetch('/upload-document', {
                    method: 'POST',
                    body: formData
                });

                const result = await response.json();

                if (response.ok) {
                    showMessage('uploadMessage', 'success', '✓ ' + result.message);
                    removeFile();
                    document.getElementById('title').value = '';
                } else {
                    showMessage('uploadMessage', 'error', '✕ ' + (result.message || 'アップロードに失敗しました'));
                    uploadBtn.disabled = false;
                }
            } catch (error) {
                showMessage('uploadMessage', 'error', '✕ エラーが発生しました: ' + error.message);
                uploadBtn.disabled = false;
            } finally {
                loader.classList.remove('show');
            }
        });
        }); // DOMContentLoaded終了

        // グローバル関数定義（onclick属性から呼ばれるため）

        function getFileIcon(filename) {
            const ext = filename.split('.').pop().toLowerCase();
            if (ext === 'pdf') return '📕';
            if (ext === 'xlsx' || ext === 'xls') return '📊';
            if (ext === 'txt') return '📄';
            return '📁';
        }

        function removeFile() {
            selectedFile = null;
            document.getElementById('selectedFile').classList.remove('show');
            document.getElementById('uploadBtn').disabled = true;
            document.getElementById('fileInput').value = '';
        }

        // ファイル一覧を読み込む
        async function loadDocuments() {
            const loader = document.getElementById('listLoader');
            const fileList = document.getElementById('fileList');

            loader.classList.add('show');
            hideMessage('listMessage');
            fileList.innerHTML = '';

            try {
                const response = await fetch('/documents');
                const result = await response.json();

                if (response.ok && result.status === 'success') {
                    if (result.documents.length === 0) {
                        fileList.innerHTML = `
                            <div class="empty-state">
                                <div class="empty-icon">📂</div>
                                <div class="empty-text">まだファイルがアップロードされていません</div>
                            </div>
                        `;
                    } else {
                        fileList.innerHTML = result.documents.map(doc => {
                            const icon = getFileIcon(doc.title);
                            const chunks = doc.chunk_count || 0;
                            const hasEmbeddings = doc.has_embeddings;
                            return `
                            <div class="file-item">
                                <div class="file-item-icon">${icon}</div>
                                <div class="file-item-info">
                                    <div class="file-item-title">${doc.title}</div>
                                    <div class="file-item-meta">
                                        <span class="file-item-badge">${doc.source_type}</span>
                                        ${chunks} チャンク${hasEmbeddings ? ' • Embedding済み' : ''}
                                    </div>
                                </div>
                                <div class="file-item-actions">
                                    ${!hasEmbeddings ? `<button class="icon-btn" onclick="generateEmbedding('${doc.source_id}', '${doc.source_type}')" title="Embedding生成">🔮</button>` : ''}
                                    <button class="icon-btn" onclick="downloadDocument('${doc.source_id}', '${doc.source_type}')" title="ダウンロード">⬇️</button>
                                    <button class="icon-btn" onclick="deleteDocument('${doc.source_id}', '${doc.source_type}')" title="削除">🗑️</button>
                                </div>
                            </div>
                        `}).join('');
                    }
                } else {
                    showMessage('listMessage', 'error', '✕ ファイル一覧の取得に失敗しました');
                }
            } catch (error) {
                showMessage('listMessage', 'error', '✕ エラーが発生しました: ' + error.message);
            } finally {
                loader.classList.remove('show');
            }
        }

        // 全てのEmbedding生成
        async function generateAllEmbeddings() {
            if (!confirm('全てのEmbedding未生成ファイルに対してEmbeddingを生成しますか？\\n\\n※ファイル数によっては時間がかかる場合があります')) return;

            const loader = document.getElementById('listLoader');
            loader.classList.add('show');
            hideMessage('listMessage');

            try {
                const response = await fetch('/generate-embeddings', {
                    method: 'POST'
                });

                const result = await response.json();

                if (response.ok) {
                    showMessage('listMessage', 'success', '✓ ' + result.message);
                    loadDocuments();
                } else {
                    showMessage('listMessage', 'error', '✕ ' + (result.message || 'Embedding生成に失敗しました'));
                }
            } catch (error) {
                showMessage('listMessage', 'error', '✕ エラーが発生しました: ' + error.message);
            } finally {
                loader.classList.remove('show');
            }
        }

        // 個別ファイルのEmbedding生成
        async function generateEmbedding(sourceId, sourceType) {
            if (!confirm('このファイルのEmbeddingを生成しますか？')) return;

            const loader = document.getElementById('listLoader');
            loader.classList.add('show');
            hideMessage('listMessage');

            try {
                const response = await fetch(`/generate-embeddings/${sourceId}?source_type=${sourceType}`, {
                    method: 'POST'
                });

                const result = await response.json();

                if (response.ok) {
                    showMessage('listMessage', 'success', '✓ ' + result.message);
                    loadDocuments();
                } else {
                    showMessage('listMessage', 'error', '✕ ' + (result.message || 'Embedding生成に失敗しました'));
                }
            } catch (error) {
                showMessage('listMessage', 'error', '✕ エラーが発生しました: ' + error.message);
            } finally {
                loader.classList.remove('show');
            }
        }

        // ファイルダウンロード
        function downloadDocument(sourceId, sourceType) {
            const url = `/download-document/${sourceId}?source_type=${sourceType}`;
            window.location.href = url;
        }

        // ファイル削除
        async function deleteDocument(sourceId, sourceType) {
            if (!confirm('このファイルを削除しますか？')) return;

            try {
                const response = await fetch(`/delete-document/${sourceId}?source_type=${sourceType}`, {
                    method: 'DELETE'
                });

                const result = await response.json();

                if (response.ok) {
                    showMessage('listMessage', 'success', '✓ ' + result.message);
                    loadDocuments();
                } else {
                    showMessage('listMessage', 'error', '✕ ' + (result.message || '削除に失敗しました'));
                }
            } catch (error) {
                showMessage('listMessage', 'error', '✕ エラーが発生しました: ' + error.message);
            }
        }

        // メッセージ表示ヘルパー
        function showMessage(elementId, type, text) {
            const message = document.getElementById(elementId);
            message.className = 'message show ' + type;
            message.innerHTML = `
                <span class="message-icon">${type === 'success' ? '✓' : '⚠'}</span>
                <span>${text}</span>
            `;
        }

        function hideMessage(elementId) {
            const message = document.getElementById(elementId);
            message.classList.remove('show');
        }
    </script>
</body>
</html>
    """
    from flask import make_response
    response = make_response(html)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    # XSS対策: Content Security Policy
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; img-src 'self' data:; font-src 'self'; connect-src 'self'"
    # クリックジャッキング対策
    response.headers['X-Frame-Options'] = 'DENY'
    # MIME type sniffing対策
    response.headers['X-Content-Type-Options'] = 'nosniff'
    # XSS保護
    response.headers['X-XSS-Protection'] = '1; mode=block'
    return response


@app.route("/upload-document", methods=["POST"])
def upload_document_public():
    """ファイルをアップロードしてRAGに追加（誰でもアクセス可能）"""
    try:
        # クライアントIPアドレスを取得（プロキシ経由の場合も考慮）
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        if ',' in client_ip:
            client_ip = client_ip.split(',')[0].strip()

        # レート制限チェック
        if not check_upload_rate_limit(client_ip):
            logger.warning(f"アップロードレート制限超過: IP={client_ip}")
            return jsonify({
                "status": "error",
                "message": f"アップロード回数の上限に達しました。1時間あたり{Config.UPLOAD_RATE_LIMIT_PER_HOUR}回までです。"
            }), 429

        # RAGサービスの状態確認
        if not rag_service or not rag_service.is_enabled:
            return jsonify({
                "status": "error",
                "message": "RAGサービスが無効です。管理者に連絡してください。"
            }), 503

        # ファイルの取得
        if 'file' not in request.files:
            return jsonify({
                "status": "error",
                "message": "ファイルが指定されていません"
            }), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({
                "status": "error",
                "message": "ファイル名が空です"
            }), 400

        # ファイルサイズチェック（Flaskの自動チェックに加えて明示的にチェック）
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)

        # Railwayのタイムアウト対策: 5MB以下に制限
        max_size_bytes_strict = 5 * 1024 * 1024  # 5MB
        file_size_mb = file_size / (1024 * 1024)

        if file_size > max_size_bytes_strict:
            logger.warning(f"ファイルサイズ超過: {file_size_mb:.2f}MB (最大: 5MB)")
            return jsonify({
                "status": "error",
                "message": f"ファイルサイズが大きすぎます（{file_size_mb:.2f}MB）。最大5MBまでです。大きなファイルは分割してアップロードしてください。"
            }), 413

        max_size_bytes = Config.MAX_FILE_SIZE_MB * 1024 * 1024
        if file_size > max_size_bytes:
            logger.warning(f"ファイルサイズ超過: {file_size} bytes (最大: {max_size_bytes} bytes)")
            return jsonify({
                "status": "error",
                "message": f"ファイルサイズが大きすぎます。最大{Config.MAX_FILE_SIZE_MB}MBまでです。"
            }), 413

        # ファイルタイトルの取得（オプション）
        title = request.form.get('title', file.filename)

        # ファイルタイプの判定
        filename = file.filename.lower()

        # ファイルの内容を読み込み
        content = None

        if filename.endswith('.pdf'):
            # PDFファイルの処理
            import io
            try:
                import pdfplumber
                pdf_content = file.read()
                pdf_file = io.BytesIO(pdf_content)
                text_parts = []

                with pdfplumber.open(pdf_file) as pdf:
                    for page_num, page in enumerate(pdf.pages, 1):
                        text = page.extract_text()
                        if text:
                            text_parts.append(f"=== ページ {page_num} ===\n{text}")

                content = "\n\n".join(text_parts)
                logger.info(f"PDFファイルを解析しました: {filename}, {len(pdf.pages)}ページ")

            except Exception as e:
                error_msg = safe_error_message(e, "PDFファイルの解析に失敗しました")
                return jsonify({
                    "status": "error",
                    "message": error_msg
                }), 500

        elif filename.endswith(('.xlsx', '.xls')):
            # Excelファイルの処理
            import io
            try:
                import openpyxl
                excel_content = file.read()
                excel_file = io.BytesIO(excel_content)
                workbook = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)

                text_parts = []

                sheet_count = 0
                for sheet_name in workbook.sheetnames:
                    sheet_count += 1
                    if sheet_count > Config.EXCEL_MAX_SHEETS:
                        text_parts.append(f"... (残り{len(workbook.sheetnames) - Config.EXCEL_MAX_SHEETS}シートは省略されました)")
                        break
                    worksheet = workbook[sheet_name]
                    sheet_text = [f"=== シート: {sheet_name} ==="]

                    # ヘッダー行
                    headers = []
                    for cell in worksheet[1]:
                        if cell.value:
                            headers.append(str(cell.value))

                    if headers:
                        sheet_text.append(f"列: {', '.join(headers)}")

                    # データ行（最大行数まで - タイムアウト対策）
                    row_count = 0
                    # max_rowで明示的に行数を制限（read_only=Trueでも確実に高速化）
                    max_row_limit = min(Config.EXCEL_MAX_ROWS_PER_SHEET + 1, worksheet.max_row)  # ヘッダー+設定値

                    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_row=max_row_limit, values_only=True), 2):
                        row_values = [str(val) if val is not None else "" for val in row]
                        if any(row_values):
                            row_text = " | ".join([f"{h}={v}" for h, v in zip(headers, row_values) if v])
                            if row_text:
                                sheet_text.append(f"行{row_num}: {row_text}")
                                row_count += 1

                    # 省略された行がある場合は通知
                    if worksheet.max_row > max_row_limit:
                        omitted_rows = worksheet.max_row - max_row_limit
                        sheet_text.append(f"... (残り{omitted_rows}行は省略されました。最大{Config.EXCEL_MAX_ROWS_PER_SHEET}行まで処理)")

                    text_parts.append("\n".join(sheet_text))

                content = "\n\n".join(text_parts)
                logger.info(f"Excelファイルを解析しました: {filename}, {len(workbook.sheetnames)}シート")

            except Exception as e:
                error_msg = safe_error_message(e, "Excelファイルの解析に失敗しました")
                return jsonify({
                    "status": "error",
                    "message": error_msg
                }), 500

        elif filename.endswith('.txt'):
            # テキストファイルの処理
            try:
                content = file.read().decode('utf-8', errors='ignore')
                logger.info(f"テキストファイルを読み込みました: {filename}")
            except Exception as e:
                error_msg = safe_error_message(e, "テキストファイルの読み込みに失敗しました")
                return jsonify({
                    "status": "error",
                    "message": error_msg
                }), 500
        else:
            return jsonify({
                "status": "error",
                "message": "対応していないファイル形式です。PDF、Excel、またはテキストファイルをアップロードしてください。"
            }), 400

        # 内容が空でないかチェック
        if not content or not content.strip():
            return jsonify({
                "status": "error",
                "message": "ファイルの内容が空です"
            }), 400

        # 内容のサイズチェック（大きすぎる場合は警告）
        content_size_mb = len(content.encode('utf-8')) / (1024 * 1024)
        if content_size_mb > 5:
            logger.warning(f"大きなファイルをアップロード中: {content_size_mb:.2f}MB")

        # RAGサービスに追加
        try:
            import hashlib
            file_hash = hashlib.md5(content.encode()).hexdigest()

            logger.info(f"RAGへの追加を開始: {title}, サイズ: {content_size_mb:.2f}MB")

            # タイムアウト対策: 小さいファイル（1MB未満）は自動生成、大きいファイルは後で生成
            # 大きなファイルの場合、アップロード後に /generate-embeddings エンドポイントで生成してください
            auto_generate_threshold_mb = 1.0  # 1MB未満は自動生成
            generate_embeddings_now = content_size_mb < auto_generate_threshold_mb

            if generate_embeddings_now:
                logger.info(f"Embedding生成を自動実行します（サイズ: {content_size_mb:.2f}MB < {auto_generate_threshold_mb}MB）")
            else:
                logger.info(f"Embedding生成は後で実行します（サイズ: {content_size_mb:.2f}MB >= {auto_generate_threshold_mb}MB、タイムアウト対策）")

            success = rag_service.add_document(
                source_type="upload",
                source_id=f"upload_{file_hash}",
                title=title,
                content=content,
                metadata={
                    "filename": file.filename,
                    "uploaded_at": datetime.now().isoformat(),
                    "file_type": filename.split('.')[-1],
                    "content_size_mb": round(content_size_mb, 2),
                    "embeddings_generated": generate_embeddings_now
                },
                generate_embeddings=generate_embeddings_now
            )

            if not success:
                return jsonify({
                    "status": "error",
                    "message": "ファイルの追加に失敗しました"
                }), 500

            logger.info(f"ファイルをRAGに追加しました: {title}")

            return jsonify({
                "status": "success",
                "message": f"ファイル '{title}' をRAGに追加しました。「Embedding生成」ボタンをクリックして検索可能にしてください。",
                "filename": file.filename,
                "title": title,
                "content_length": len(content),
                "content_size_mb": round(content_size_mb, 2),
                "embeddings_generated": False,
                "next_step": "画面の「Embedding生成」ボタンをクリックしてください"
            })

        except Exception as e:
            logger.error(f"RAGへの追加エラー: {e}", exc_info=True)
            error_msg = safe_error_message(e, "ファイルの処理中にエラーが発生しました")
            return jsonify({
                "status": "error",
                "message": error_msg
            }), 500

    except Exception as e:
        logger.error("ファイルアップロードAPIでエラーが発生しました", error=str(e), exc_info=True)
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/admin/upload-document", methods=["POST"])
@require_admin
def upload_document_admin():
    """ファイルをアップロードしてRAGに追加（管理者のみ - 認証あり）"""
    # 公開エンドポイントと同じ処理を使用
    return upload_document_public()


@app.route("/admin/stats", methods=["GET"])
@require_admin
def get_stats():
    """統計情報の取得（管理者のみ）"""
    try:
        qa_stats = qa_service.get_stats()
        
        # 統計を結合
        combined_stats = qa_stats.to_dict()
        combined_stats["total_flows"] = len(flow_service.flows)
        
        return jsonify(combined_stats)
    except Exception as e:
        logger.error("統計情報の取得に失敗しました", error=str(e))
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/admin/auto-reload/status", methods=["GET"])
@require_admin
def get_auto_reload_status():
    """自動リロードの状態確認"""
    try:
        return jsonify({
            "status": "success",
            "auto_reload_active": True,
            "last_reload": time.time(),
            "next_reload_in_seconds": 900,  # 15分後（API制限を考慮）
            "message": "自動リロードが動作中です"
        })
    except Exception as e:
        logger.error("自動リロード状態の取得に失敗しました", error=str(e))
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/admin/rag-diagnostic", methods=["GET"])
@require_admin
def rag_diagnostic():
    """RAGサービスの診断（Railway環境で問題を特定するため）"""
    try:
        # 環境変数の確認
        env_check = {
            "GEMINI_API_KEY": bool(os.getenv('GEMINI_API_KEY')),
            "DATABASE_URL": bool(os.getenv('DATABASE_URL')),
            "RAG_LIGHTWEIGHT_MODE": os.getenv('RAG_LIGHTWEIGHT_MODE', 'false'),
            "SIMILARITY_THRESHOLD": os.getenv('SIMILARITY_THRESHOLD', '0.15'),
            "EMBEDDING_MODEL": os.getenv('EMBEDDING_MODEL', 'sentence-transformers/all-MiniLM-L6-v2'),
        }

        # RAGServiceの状態
        rag_status = {
            "is_enabled": rag_service.is_enabled,
            "db_pool_initialized": rag_service.db_pool is not None,
            "embedding_model_loaded": rag_service.embedding_model is not None,
            "gemini_model_initialized": rag_service.gemini_model is not None,
        }

        # 依存関係の確認
        dependencies = {
            "sentence_transformers": False,
            "numpy": False,
        }

        try:
            import sentence_transformers
            dependencies["sentence_transformers"] = True
        except ImportError:
            pass

        try:
            import numpy
            dependencies["numpy"] = True
        except ImportError:
            pass

        # テストクエリで検索（Embeddingモデルがロードされている場合のみ）
        test_search_results = None
        if rag_service.embedding_model and rag_service.db_pool:
            try:
                results = rag_service.search_similar_documents("配送にはどれくらい時間がかかる？", limit=10)
                test_search_results = {
                    "total_results": len(results),
                    "top_3": [
                        {
                            "title": r.get('title', 'N/A'),
                            "similarity": round(r.get('similarity', 0), 4)
                        }
                        for r in results[:3]
                    ]
                }
            except Exception as search_error:
                test_search_results = {"error": str(search_error)}

        # 診断結果のサマリー
        diagnosis = "unknown"
        recommendation = ""

        if rag_status["is_enabled"] and rag_status["embedding_model_loaded"] and rag_status["db_pool_initialized"]:
            diagnosis = "healthy"
            recommendation = "RAGシステムは正常に動作しています"
        elif rag_status["is_enabled"] and not rag_status["embedding_model_loaded"]:
            diagnosis = "embedding_model_not_loaded"
            if env_check["RAG_LIGHTWEIGHT_MODE"].lower() == 'true':
                recommendation = "RAG_LIGHTWEIGHT_MODE=true が設定されています。環境変数を削除してください。"
            elif not dependencies["sentence_transformers"]:
                recommendation = "sentence-transformersがインストールされていません。依存関係を確認してください。"
            else:
                recommendation = "Embeddingモデルのロードに失敗しています。メモリ不足またはタイムアウトの可能性があります。"
        elif not rag_status["db_pool_initialized"]:
            diagnosis = "database_not_connected"
            recommendation = "データベース接続に失敗しています。DATABASE_URLを確認してください。"
        else:
            diagnosis = "rag_disabled"
            recommendation = "RAG機能が無効化されています。ログを確認してください。"

        return jsonify({
            "status": "success",
            "diagnosis": diagnosis,
            "recommendation": recommendation,
            "environment_variables": env_check,
            "rag_status": rag_status,
            "dependencies": dependencies,
            "test_search_results": test_search_results,
            "timestamp": datetime.utcnow().isoformat()
        })

    except Exception as e:
        logger.error("RAG診断に失敗しました", error=str(e))
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/admin/authenticated-users", methods=["GET"])
@require_admin
def get_authenticated_users():
    """認証済みユーザー一覧を取得"""
    try:
        from .optimized_auth_flow import OptimizedAuthFlow
        
        auth_flow = OptimizedAuthFlow()
        stats = auth_flow.get_stats()
        
        # 認証済みユーザーの詳細情報を取得
        authenticated_users = []
        for user_id, auth_info in auth_flow.authenticated_users.items():
            authenticated_users.append({
                "user_id": hash_user_id(user_id),
                "store_code": auth_info.get('store_code'),
                "staff_id": auth_info.get('staff_id'),
                "store_name": auth_info.get('store_name'),
                "staff_name": auth_info.get('staff_name'),
                "auth_time": auth_info.get('auth_time')
            })
        
        return jsonify({
            "status": "success",
            "total_authenticated": stats['total_authenticated'],
            "authenticated_users": authenticated_users,
            "cache_valid": stats.get('cache_valid', False),
            "last_cache_update": stats.get('last_cache_update', 0),
            "last_updated": stats['last_updated']
        })
        
    except Exception as e:
        logger.error("認証済みユーザー一覧の取得に失敗しました", error=str(e))
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/admin/force-cache-update", methods=["POST"])
@require_admin
def force_cache_update():
    """キャッシュを強制更新"""
    try:
        from .optimized_auth_flow import OptimizedAuthFlow
        
        auth_flow = OptimizedAuthFlow()
        auth_flow.force_cache_update()
        
        logger.info("管理者によるキャッシュ強制更新が完了しました")
        return jsonify({
            "status": "success",
            "message": "キャッシュを強制更新しました",
            "cache_valid": auth_flow._is_cache_valid(),
            "last_cache_update": auth_flow.last_cache_update
        })
        
    except Exception as e:
        logger.error("キャッシュ強制更新に失敗しました", error=str(e))
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/admin/check-all-users-status", methods=["POST"])
@require_admin
def check_all_users_status():
    """全認証済みユーザーのステータスを即座にチェック"""
    try:
        from .optimized_auth_flow import OptimizedAuthFlow
        
        auth_flow = OptimizedAuthFlow()
        result = auth_flow.check_all_users_status()
        
        if result:
            logger.info("管理者による全ユーザーステータスチェックが完了しました")
            return jsonify({
                "status": "success",
                "message": "全ユーザーのステータスチェックが完了しました",
                "total_checked": result['total_checked'],
                "deauthenticated_count": result['deauthenticated_count'],
                "deauthenticated_users": [hash_user_id(uid) for uid in result['deauthenticated_users']]
            })
        else:
            return jsonify({
                "status": "error",
                "message": "全ユーザーのステータスチェックに失敗しました"
            }), 500
        
    except Exception as e:
        logger.error("全ユーザーステータスチェックに失敗しました", error=str(e))
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/admin/deauthenticate", methods=["POST"])
@require_admin
def deauthenticate_user():
    """ユーザーの認証を取り消す"""
    try:
        from .optimized_auth_flow import OptimizedAuthFlow

        data = request.get_json()
        if not data or 'user_id' not in data:
            return jsonify({
                "status": "error",
                "message": "user_idが必要です"
            }), 400

        user_id = data['user_id']
        auth_flow = OptimizedAuthFlow()

        # 認証を取り消し
        success = auth_flow.deauthenticate_user(user_id)

        if success:
            logger.info("管理者による認証取り消しが完了しました",
                       user_id=hash_user_id(user_id))
            return jsonify({
                "status": "success",
                "message": f"ユーザー {hash_user_id(user_id)} の認証を取り消しました"
            })
        else:
            return jsonify({
                "status": "error",
                "message": f"ユーザー {hash_user_id(user_id)} の認証取り消しに失敗しました"
            }), 400

    except Exception as e:
        logger.error("認証取り消しに失敗しました", error=str(e))
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/admin/auth-db-status", methods=["GET"])
@require_admin
def get_auth_db_status():
    """認証データベースの状態を確認"""
    try:
        from .auth_service import AuthService
        from .auth_db_service import AuthDBService

        # AuthServiceの状態
        auth_service = AuthService()

        # AuthDBServiceの直接チェック
        auth_db = AuthDBService()

        status = {
            "database_url_set": bool(os.getenv('DATABASE_URL')),
            "auth_db_enabled": auth_db.is_enabled,
            "auth_db_health": auth_db.health_check() if auth_db.is_enabled else False,
            "auth_service_db_enabled": auth_service.auth_db.is_enabled,
        }

        # データベースからユーザー数を取得
        if auth_db.is_enabled:
            try:
                users = auth_db.get_all_authenticated_users()
                status["total_users_in_db"] = len(users)
                status["users"] = [
                    {
                        "line_user_id": hash_user_id(user['line_user_id']),
                        "store_code": user['store_code'],
                        "staff_id": user['staff_id'],
                        "staff_name": user['staff_name'],
                        "auth_time": str(user['auth_time'])
                    }
                    for user in users[:10]  # 最大10件
                ]
            except Exception as e:
                status["db_query_error"] = str(e)
                status["total_users_in_db"] = "error"

        return jsonify({
            "status": "success",
            **status
        })

    except Exception as e:
        logger.error("認証DB状態の取得に失敗しました", error=str(e))
        import traceback
        return jsonify({
            "status": "error",
            "message": str(e),
            "trace": traceback.format_exc()
        }), 500


@app.errorhandler(400)
def bad_request(error):
    return jsonify({"error": "Bad Request", "message": error.description}), 400


@app.errorhandler(401)
def unauthorized(error):
    return jsonify({"error": "Unauthorized", "message": error.description}), 401


@app.errorhandler(403)
def forbidden(error):
    return jsonify({"error": "Forbidden", "message": error.description}), 403


@app.errorhandler(500)
def internal_error(error):
    return jsonify({"error": "Internal Server Error", "message": "内部エラーが発生しました"}), 500


def main():
    """アプリケーションの起動"""
    try:
        # サービスを初期化
        logger.info("サービスを初期化しています...")
        initialize_services()
        logger.info("サービス初期化完了")
        
        port = int(os.environ.get("PORT", 5000))
        logger.info(f"アプリケーションを起動します (ポート: {port})")
        app.run(host="0.0.0.0", port=port, debug=False)
    except Exception as e:
        logger.error("アプリケーションの起動に失敗しました", error=str(e))
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()
