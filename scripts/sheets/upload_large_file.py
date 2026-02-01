#!/usr/bin/env python3
"""
大きなファイルをRAGシステムに直接アップロードするスクリプト
Webアップロードでタイムアウトする場合に使用
"""

import sys
import os
from datetime import datetime

# 環境変数の読み込み
from dotenv import load_dotenv
load_dotenv()

def upload_large_file(file_path: str, title: str = None):
    """大きなファイルをRAGに追加"""

    # ファイルの存在確認
    if not os.path.exists(file_path):
        print(f"❌ ファイルが見つかりません: {file_path}")
        return False

    # ファイルサイズ確認
    file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
    print(f"📁 ファイルサイズ: {file_size_mb:.2f}MB")

    # タイトルが指定されていない場合はファイル名を使用
    if not title:
        title = os.path.basename(file_path)

    print(f"📝 タイトル: {title}")
    print(f"🔄 処理を開始します...")

    # RAGServiceをインポート
    try:
        from line_qa_system.rag_service import RAGService
        from line_qa_system.config import Config

        print("✅ RAGServiceをインポートしました")

        # RAGServiceの初期化
        rag_service = RAGService()

        if not rag_service.is_enabled:
            print("❌ RAGサービスが無効です")
            return False

        print("✅ RAGサービスを初期化しました")

    except Exception as e:
        print(f"❌ RAGServiceのインポートに失敗: {e}")
        return False

    # ファイルの内容を読み込み
    try:
        filename = file_path.lower()

        if filename.endswith('.pdf'):
            # PDFファイルの処理
            print("📄 PDFファイルを解析中...")
            import io
            import pdfplumber

            with open(file_path, 'rb') as f:
                pdf_content = f.read()

            pdf_file = io.BytesIO(pdf_content)
            text_parts = []

            with pdfplumber.open(pdf_file) as pdf:
                print(f"📄 {len(pdf.pages)}ページを処理中...")
                for page_num, page in enumerate(pdf.pages, 1):
                    if page_num % 10 == 0:
                        print(f"  処理中: {page_num}/{len(pdf.pages)}ページ")
                    text = page.extract_text()
                    if text:
                        text_parts.append(f"=== ページ {page_num} ===\n{text}")

            content = "\n\n".join(text_parts)
            print(f"✅ PDFファイルを解析しました: {len(pdf.pages)}ページ")

        elif filename.endswith(('.xlsx', '.xls')):
            # Excelファイルの処理
            print("📊 Excelファイルを解析中...")
            import io
            import openpyxl

            with open(file_path, 'rb') as f:
                excel_content = f.read()

            excel_file = io.BytesIO(excel_content)
            workbook = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)

            text_parts = []
            MAX_ROWS_PER_SHEET = 1000  # より多くの行を処理
            MAX_SHEETS = 20  # より多くのシートを処理

            sheet_count = 0
            print(f"📊 {len(workbook.sheetnames)}シートを処理中...")

            for sheet_name in workbook.sheetnames:
                sheet_count += 1
                if sheet_count > MAX_SHEETS:
                    text_parts.append(f"... (残り{len(workbook.sheetnames) - MAX_SHEETS}シートは省略されました)")
                    break

                print(f"  シート {sheet_count}/{min(len(workbook.sheetnames), MAX_SHEETS)}: {sheet_name}")

                worksheet = workbook[sheet_name]
                sheet_text = [f"=== シート: {sheet_name} ==="]

                # ヘッダー行
                headers = []
                for cell in worksheet[1]:
                    if cell.value:
                        headers.append(str(cell.value))

                if headers:
                    sheet_text.append(f"列: {', '.join(headers)}")

                # データ行
                row_count = 0
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                    if row_count >= MAX_ROWS_PER_SHEET:
                        sheet_text.append(f"... (残りの行は省略されました。最大{MAX_ROWS_PER_SHEET}行まで)")
                        break

                    row_values = [str(val) if val is not None else "" for val in row]
                    if any(row_values):
                        row_text = " | ".join([f"{h}={v}" for h, v in zip(headers, row_values) if v])
                        if row_text:
                            sheet_text.append(f"行{row_num}: {row_text}")
                            row_count += 1

                text_parts.append("\n".join(sheet_text))

            content = "\n\n".join(text_parts)
            print(f"✅ Excelファイルを解析しました: {min(len(workbook.sheetnames), MAX_SHEETS)}シート")

        elif filename.endswith('.txt'):
            # テキストファイルの処理
            print("📝 テキストファイルを読み込み中...")
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            print(f"✅ テキストファイルを読み込みました")

        else:
            print(f"❌ 対応していないファイル形式です: {filename}")
            return False

        # 内容が空でないかチェック
        if not content or not content.strip():
            print("❌ ファイルの内容が空です")
            return False

        content_size_mb = len(content.encode('utf-8')) / (1024 * 1024)
        print(f"📊 解析後のサイズ: {content_size_mb:.2f}MB")

    except Exception as e:
        print(f"❌ ファイルの解析に失敗: {e}")
        import traceback
        traceback.print_exc()
        return False

    # RAGサービスに追加
    try:
        import hashlib
        file_hash = hashlib.md5(content.encode()).hexdigest()

        print("💾 RAGデータベースに保存中...")
        print("   （Embeddingは後で生成されます）")

        success = rag_service.add_document(
            source_type="upload",
            source_id=f"upload_{file_hash}",
            title=title,
            content=content,
            metadata={
                "filename": os.path.basename(file_path),
                "uploaded_at": datetime.now().isoformat(),
                "file_type": file_path.split('.')[-1],
                "content_size_mb": round(content_size_mb, 2),
                "embeddings_generated": False,
                "upload_method": "direct_script"
            },
            generate_embeddings=False  # Embeddingは後で生成
        )

        if success:
            print("✅ ファイルをRAGに追加しました！")
            print(f"   タイトル: {title}")
            print(f"   ファイル名: {os.path.basename(file_path)}")
            print(f"   サイズ: {content_size_mb:.2f}MB")
            print("")
            print("📌 注意: Embeddingは後で生成されます")
            print("   検索機能を使うにはEmbeddingの生成が必要です")
            return True
        else:
            print("❌ RAGへの追加に失敗しました")
            return False

    except Exception as e:
        print(f"❌ RAGへの追加エラー: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使い方: python3 upload_large_file.py <ファイルパス> [タイトル]")
        print("")
        print("例:")
        print('  python3 upload_large_file.py "/path/to/file.xlsx" "営業関連シート"')
        print('  python3 upload_large_file.py "/path/to/file.pdf"')
        sys.exit(1)

    file_path = sys.argv[1]
    title = sys.argv[2] if len(sys.argv) > 2 else None

    print("=" * 60)
    print("📤 大きなファイルのアップロードスクリプト")
    print("=" * 60)
    print("")

    success = upload_large_file(file_path, title)

    print("")
    print("=" * 60)
    if success:
        print("✅ アップロード成功")
    else:
        print("❌ アップロード失敗")
    print("=" * 60)

    sys.exit(0 if success else 1)
