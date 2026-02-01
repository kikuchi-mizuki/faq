#!/usr/bin/env python3
"""
Excelファイルを直接RAGに追加
"""

import os
import sys
import json
import base64
import io
from dotenv import load_dotenv
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
import openpyxl

# 環境変数の読み込み
load_dotenv()

# パスを追加
sys.path.insert(0, os.path.dirname(__file__))

def add_excel_to_rag():
    """Excelファイルを直接RAGに追加"""

    print("=" * 60)
    print("Excelファイルを直接RAGに追加")
    print("=" * 60)

    # Excelファイルの情報
    excel_file_id = "1QvLDETWheVu8QYTZaSJLI1jy__70-AbV"
    excel_file_name = "菊池さん共有_営業関連シート.xlsx"

    # 認証情報の取得
    print("\n🔧 認証情報を取得中...")
    service_account_json_raw = os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON')

    if service_account_json_raw.endswith('.json'):
        with open(service_account_json_raw, 'r') as f:
            credentials_info = json.load(f)
    else:
        service_account_json = base64.b64decode(service_account_json_raw).decode('utf-8')
        credentials_info = json.loads(service_account_json)

    credentials = Credentials.from_service_account_info(
        credentials_info,
        scopes=[
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive',
            'https://www.googleapis.com/auth/documents'
        ]
    )

    print("✅ 認証情報を取得しました")

    # Google Drive APIサービスを初期化
    drive_service = build('drive', 'v3', credentials=credentials)

    # Excelファイルをダウンロード
    print(f"\n📊 {excel_file_name} をダウンロード中...")
    request = drive_service.files().get_media(fileId=excel_file_id)
    excel_content = request.execute()
    print(f"   ✅ ダウンロード完了（{len(excel_content)}バイト）")

    # openpyxlで読み込み
    print("\n📖 Excelファイルを解析中...")
    excel_file = io.BytesIO(excel_content)
    workbook = openpyxl.load_workbook(excel_file, data_only=True)

    text_parts = []
    for sheet_name in workbook.sheetnames:
        print(f"   - シート '{sheet_name}' を処理中...")
        worksheet = workbook[sheet_name]
        sheet_text = [f"=== シート: {sheet_name} ==="]

        # ヘッダー行の取得
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value))

        if headers:
            sheet_text.append(f"列: {', '.join(headers)}")

        # データ行の処理
        row_count = 0
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
            row_values = [str(val) if val is not None else "" for val in row]
            if any(row_values):
                row_text = " | ".join([f"{h}={v}" for h, v in zip(headers, row_values) if v])
                if row_text:
                    sheet_text.append(f"行{row_num}: {row_text}")
                    row_count += 1

        print(f"      → {row_count}行のデータを抽出")
        text_parts.append("\n".join(sheet_text))

    content = "\n\n".join(text_parts)
    print(f"\n✅ 合計{len(workbook.sheetnames)}シートから{len(content)}文字を抽出しました")

    # RAGサービスに追加
    print("\n📚 RAGサービスに追加中...")
    from line_qa_system.rag_service import RAGService

    rag_service = RAGService()

    if not rag_service.is_enabled:
        print("❌ RAGサービスが有効になっていません")
        return

    rag_service.add_document(
        source_type="google_drive",
        source_id=excel_file_id,
        title=excel_file_name,
        content=content,
        metadata={
            "file_id": excel_file_id,
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "modified_time": "2026-01-18T09:41:48.000Z",
            "collected_at": "manual_collection_excel"
        }
    )

    print(f"✅ RAGに追加しました")
    print("\n" + "=" * 60)
    print("完了")
    print("=" * 60)

if __name__ == "__main__":
    add_excel_to_rag()
